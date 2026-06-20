"""tie_out - numeric safety gate for the MRO native rewrite (Phase 4 = final form).

The build emits every formula with NO cached `<v>`, so headless LibreOffice must
compute results on load. This harness recomputes a built workbook with `soffice` and
checks two relocation-proof invariants against the frozen v4.33 oracle in
`qa/gold/baseline.json`:

  * Invariant B - accessor-target spot checks (the load-bearing model gate). The
    workbook ships ZERO defined names now, so there is no definedName table to read.
    Instead `qa/name_map.NAME_TO_ACCESSOR` maps each of the 86 legacy v4.33 names to
    its producer accessor's `'Sheet'!Cell` ref; this harness reads the recomputed value
    at that cell and asserts it equals `baseline.json["defined_names"][name]` within
    tol. The accessors are the SAME closures the consumer sheets use, so a wrong
    captured row fails here AND breaks the live model identically - the gate cannot
    agree with itself against a broken build. (Through Phase 3 this was fed by the
    workbook's bridge definedNames; same values, different source.)

  * Invariant A - engine value-multiset (relocation-proof tripwire). One workbook-wide
    multiset of every numeric cell over the ENGINE tabs only (the 9 calc sheets), so
    a number may move between engine sheets and still tie out as long as the SET of
    computed values is unchanged. Excludes the 3 raw data dumps and the pure link/label
    tabs (summary / guide / inputs / outputs / sources) that merely re-present figures.

Anti-laundering baseline regen: `baseline.json["defined_names"]` stays FROZEN as the
v4.33 oracle (never rewritten). `regen-baseline` only (re)writes the `engine_multiset`
block, and it refuses to do so unless Invariant B already holds against the frozen
oracle - so a regression can never be laundered into a fresh baseline.

CLI:
  python qa/tie_out.py snapshot       <built.xlsx> <out.json>
  python qa/tie_out.py compare        <gold.json>  <built.xlsx> [--tol 1.0] [--invariant-a {fail,warn,off}]
  python qa/tie_out.py regen-baseline <gold.json>  <built.xlsx> [--tol 1.0]
"""
from __future__ import annotations

import argparse
import json
import re
import subprocess
import sys
import tempfile
from pathlib import Path

import openpyxl

# name_map sets up sys.path for workbook_mro / workbook_core and imports the producer
# accessors; importing it here gives us NAME_TO_ACCESSOR (Invariant B's name table).
_HERE = Path(__file__).resolve()
sys.path.insert(0, str(_HERE.parent))                  # qa/ -> name_map
from name_map import NAME_TO_ACCESSOR, coverage_report  # noqa: E402

SOFFICE = "/opt/homebrew/bin/soffice"

# The 9 engine (calc) tabs whose numeric cells form Invariant A's value-multiset.
# Excludes the 3 raw data dumps and the pure link/label tabs (summary/guide/inputs/
# outputs/sources) that re-present numbers produced elsewhere.
_ENGINE_TABS = (
    "Reconciliation", "Services", "Depot Ship Repair", "OP-5 Navy Top-Down",
    "MSC SCN USCG Top-Down", "TAM Bridge", "Non-Public-NSY Bridge", "SAM Build",
    "Scope Reconciliation", "z_ChartData",
)

_A1 = re.compile(r"^[A-Z]{1,3}[0-9]+$")


def recompute(xlsx: Path) -> Path:
    """Convert xlsx -> xlsx through headless LibreOffice, forcing formula compute.

    Returns the path to the recomputed copy (in a fresh temp dir). Uses a private
    UserInstallation so it runs even if a desktop LibreOffice is open.
    """
    tmp = Path(tempfile.mkdtemp(prefix="tieout_"))
    profile = Path(tempfile.mkdtemp(prefix="loprof_"))
    cmd = [
        SOFFICE, "--headless", "--calc",
        f"-env:UserInstallation=file://{profile}",
        "--convert-to", "xlsx:Calc MS Excel 2007 XML",
        "--outdir", str(tmp), str(xlsx),
    ]
    proc = subprocess.run(cmd, capture_output=True, text=True, timeout=1800)
    out = tmp / (xlsx.stem + ".xlsx")
    if not out.exists():
        raise RuntimeError(
            f"soffice did not produce {out}\nstdout: {proc.stdout}\nstderr: {proc.stderr}")
    return out


def _ref_target_cell(target: str):
    """('Sheet'!$C$19) -> (sheet, 'C19'); None if not a single-cell ref."""
    if "!" not in target:
        return None
    sh, ref = target.rsplit("!", 1)
    sh = sh.strip()
    if sh.startswith("'") and sh.endswith("'"):
        sh = sh[1:-1].replace("''", "'")
    ref = ref.replace("$", "")
    if not _A1.match(ref):
        return None
    return sh, ref


def _read_cell(wb, ref: str):
    """Recomputed value at an accessor's 'Sheet'!Cell ref (or a diagnostic dict)."""
    cell = _ref_target_cell(ref)
    if cell is None:
        return {"_badref": ref}
    sh, a1 = cell
    if sh not in wb.sheetnames:
        return {"_missing_sheet": ref}
    return wb[sh][a1].value


def snapshot(xlsx: Path) -> dict:
    """Recompute `xlsx` and return the Phase-4 tie-out snapshot dict.

    accessor_values - {name: recomputed value at the name's accessor cell} for the 86
                      cell-anchored model figures (Invariant B source).
    engine_multiset - sorted numeric multiset over the 9 engine tabs (Invariant A).
    """
    rc = recompute(xlsx)
    wb = openpyxl.load_workbook(rc, data_only=True)

    accessor_values = {name: _read_cell(wb, acc()) for name, acc in NAME_TO_ACCESSOR.items()}

    engine: list[float] = []
    for sh in _ENGINE_TABS:
        if sh not in wb.sheetnames:
            continue
        for row in wb[sh].iter_rows():
            for c in row:
                v = c.value
                if isinstance(v, (int, float)) and not isinstance(v, bool):
                    engine.append(round(float(v), 2))

    return {"accessor_values": accessor_values, "engine_multiset": sorted(engine)}


def _fmt(v) -> str:
    return f"{v:,.2f}" if isinstance(v, (int, float)) else repr(v)


def _diff_multiset(gold: list, new: list, tol: float) -> list[str]:
    """Greedy nearest-match diff of two sorted numeric lists -> human messages."""
    msgs = []
    if len(gold) != len(new):
        msgs.append(f"count {len(gold)} -> {len(new)}")
    npool = sorted(new)
    used = [False] * len(npool)
    unmatched_gold = []
    for g in sorted(gold):
        hit = False
        for k, x in enumerate(npool):
            if not used[k] and abs(x - g) <= tol:
                used[k] = True
                hit = True
                break
        if not hit:
            unmatched_gold.append(g)
    unmatched_new = [npool[k] for k in range(len(npool)) if not used[k]]
    for g in unmatched_gold[:12]:
        msgs.append(f"gold value {_fmt(g)} has no match in new")
    for x in unmatched_new[:12]:
        msgs.append(f"new value {_fmt(x)} has no match in gold")
    return msgs


def _numeric_oracle(gold: dict) -> dict:
    """The numeric subset of the frozen defined-name oracle (drops the 2 date names)."""
    return {k: v for k, v in gold.get("defined_names", {}).items()
            if isinstance(v, (int, float)) and not isinstance(v, bool)}


def _invariant_b_msgs(gold: dict, new: dict, tol: float) -> list[str]:
    """Invariant B - every legacy model figure recomputes to baseline at its accessor.

    Coverage is enforced both ways: a numeric oracle name with no accessor mapping, or
    an accessor name absent from the oracle, is a failure (a silently-missed name would
    otherwise pass vacuously).
    """
    msgs: list[str] = []
    oracle = _numeric_oracle(gold)
    n_av = new.get("accessor_values", {})

    uncovered, stale = coverage_report(set(gold.get("defined_names", {})))
    if uncovered:
        msgs.append(f"INVARIANT B coverage: oracle names with no accessor: {uncovered}")
    if stale:
        msgs.append(f"INVARIANT B coverage: accessor names absent from oracle: {stale}")

    for name in sorted(set(oracle) & set(n_av)):
        gv, nv = oracle[name], n_av[name]
        if not isinstance(nv, (int, float)) or isinstance(nv, bool):
            msgs.append(f"NAME {name}: accessor cell non-numeric -> {nv!r}")
        elif abs(float(gv) - float(nv)) > tol:
            msgs.append(f"NAME {name}: {_fmt(gv)} -> {_fmt(nv)}  (Δ {_fmt(float(nv) - float(gv))})")
    return msgs


def _invariant_a_msgs(gold: dict, new: dict, tol: float):
    """Invariant A - engine value-multiset. Returns None if the oracle has no
    engine_multiset block yet (pre-regen old baseline)."""
    g = gold.get("engine_multiset")
    if g is None:
        return None
    return [f"ENGINE value-set: {m}" for m in _diff_multiset(g, new.get("engine_multiset", []), tol)]


def compare(gold: dict, new: dict, tol: float = 1.0,
            invariant_a: str = "fail") -> tuple[bool, list[str]]:
    """Diff a new snapshot against the gold oracle. Returns (ok, messages).

    Invariant B is always a hard gate. `invariant_a` in {fail, warn, off} controls the
    engine value-multiset: 'fail' (default, post-regen) folds it into the hard gate;
    'warn' prints deltas for review; 'off' skips it. If the baseline predates the
    engine_multiset block (pre-regen), Invariant A is skipped with a note - Invariant B
    is the standalone proof at that checkpoint.
    """
    msgs = _invariant_b_msgs(gold, new, tol)
    a = _invariant_a_msgs(gold, new, tol) if invariant_a != "off" else None
    if "engine_multiset" not in gold:
        print("  [Invariant A: baseline has no engine_multiset oracle yet — "
              "skipped; run regen-baseline once Invariant B is green]")
    elif invariant_a == "fail":
        msgs += a
    elif invariant_a == "warn" and a:
        print(f"  [Invariant A tripwire — {len(a)} delta(s), review only]:")
        for m in a:
            print(f"    ~ {m}")
    return (len(msgs) == 0), msgs


def regen_baseline(gold_path: Path, xlsx: Path, tol: float) -> int:
    """Self-asserting anti-laundering regen of the engine_multiset block.

    Keeps `defined_names` frozen (the v4.33 oracle); recomputes the build; asserts
    Invariant B holds against that frozen oracle; ONLY then writes the fresh
    engine_multiset. A regression in any of the 86 model figures aborts the write.
    """
    gold = json.loads(gold_path.read_text(encoding="utf-8"))
    if "defined_names" not in gold:
        print("REGEN ABORTED — gold baseline has no frozen defined_names oracle to assert against.")
        return 1

    new = snapshot(xlsx)
    b = _invariant_b_msgs(gold, new, tol)
    if b:
        print(f"REGEN ABORTED — Invariant B failed against the frozen oracle "
              f"({len(b)} discrepancies); refusing to launder a regression:")
        for m in b:
            print(f"  - {m}")
        return 1

    out = {
        "_doc": ("MRO tie-out oracle. 'defined_names' is the FROZEN v4.33 baseline "
                 "(never rewritten); Invariant B reads each via qa/name_map accessors. "
                 "'engine_multiset' is regenerated by `tie_out.py regen-baseline` and is "
                 "only written when Invariant B already holds against the frozen oracle."),
        "defined_names": gold["defined_names"],            # FROZEN — passed through verbatim
        "engine_multiset": new["engine_multiset"],
    }
    gold_path.write_text(json.dumps(out, indent=2), encoding="utf-8")
    n_named = len(_numeric_oracle(gold))
    print(f"REGEN OK — Invariant B holds for {n_named} model figures; "
          f"engine_multiset written ({len(new['engine_multiset'])} values) -> {gold_path}")
    return 0


def main(argv=None) -> int:
    ap = argparse.ArgumentParser(description=__doc__)
    sub = ap.add_subparsers(dest="cmd", required=True)
    s = sub.add_parser("snapshot")
    s.add_argument("xlsx")
    s.add_argument("out_json")
    c = sub.add_parser("compare")
    c.add_argument("gold_json")
    c.add_argument("xlsx")
    c.add_argument("--tol", type=float, default=1.0)
    c.add_argument("--invariant-a", choices=("fail", "warn", "off"), default="fail",
                   help="engine value-multiset mode (default fail; 'warn' to eyeball deltas)")
    c.add_argument("--save")
    r = sub.add_parser("regen-baseline")
    r.add_argument("gold_json")
    r.add_argument("xlsx")
    r.add_argument("--tol", type=float, default=1.0)
    args = ap.parse_args(argv)

    if args.cmd == "snapshot":
        snap = snapshot(Path(args.xlsx))
        Path(args.out_json).write_text(json.dumps(snap, indent=2), encoding="utf-8")
        n_num = sum(1 for v in snap["accessor_values"].values()
                    if isinstance(v, (int, float)) and not isinstance(v, bool))
        print(f"snapshot: {len(snap['accessor_values'])} accessor values "
              f"({n_num} numeric), {len(snap['engine_multiset'])} engine cells -> {args.out_json}")
        return 0

    if args.cmd == "regen-baseline":
        return regen_baseline(Path(args.gold_json), Path(args.xlsx), args.tol)

    gold = json.loads(Path(args.gold_json).read_text(encoding="utf-8"))
    new = snapshot(Path(args.xlsx))
    if args.save:
        Path(args.save).write_text(json.dumps(new, indent=2), encoding="utf-8")
    ok, msgs = compare(gold, new, tol=args.tol, invariant_a=args.invariant_a)
    n_named = len(_numeric_oracle(gold))
    if ok:
        print(f"TIE-OUT OK  (tol ${args.tol:g}) — Invariant B: {n_named} model figures "
              f"match at their accessors" +
              ("; Invariant A: engine value-multiset matches" if "engine_multiset" in gold else ""))
        return 0
    print(f"TIE-OUT FAILED  (tol ${args.tol:g}) — {len(msgs)} discrepancies:")
    for m in msgs:
        print(f"  - {m}")
    return 1


if __name__ == "__main__":
    sys.exit(main())
