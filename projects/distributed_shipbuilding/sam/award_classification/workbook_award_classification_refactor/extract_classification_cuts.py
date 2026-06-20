"""One-time extraction: read the manual award_classification_refactor workbook and
write one raw CSV per sheet into extracted/. Re-run only if the source changes.

Reads the MANUAL BACKUP (never the pipeline output) so re-extraction can never
pick up the pipeline's banners/sections. Cells are written verbatim as strings
(blank for empty) so identifiers keep leading zeros (Work-type ID "01", CAGE
"90099"); the sheet modules cast the known-numeric columns themselves.
"""
import csv
from pathlib import Path
import openpyxl

HERE = Path(__file__).resolve().parent
SRC = HERE.parent / "award_classification_refactor_MANUAL_BACKUP.xlsx"
OUT = HERE / "extracted"

SHEET_CSV = {
    "Taxonomy": "taxonomy",
    "Classifications (first-pass)": "classifications",
    "Vendor Context": "vendor_context",
    "DDG Top Vendors": "ddg_top_vendors",
    "Virginia Top Vendors": "virginia_top_vendors",
    "Columbia Top Vendors": "columbia_top_vendors",
}

def cellstr(v):
    if v is None:
        return ""
    if isinstance(v, float) and v.is_integer():
        # keep whole-number floats as integers (e.g. 0.0 -> "0")
        return str(int(v))
    return str(v)

wb = openpyxl.load_workbook(SRC, data_only=False)
OUT.mkdir(exist_ok=True)
for title, stem in SHEET_CSV.items():
    ws = wb[title]
    nrows, ncols = ws.max_row, ws.max_column
    with (OUT / f"{stem}.csv").open("w", encoding="utf-8", newline="") as fh:
        w = csv.writer(fh)
        for r in range(1, nrows + 1):
            w.writerow([cellstr(ws.cell(row=r, column=c).value)
                        for c in range(1, ncols + 1)])
    print(f"{title:<30s} -> extracted/{stem}.csv  ({nrows} rows x {ncols} cols)")
