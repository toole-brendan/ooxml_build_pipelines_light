"""merge_research_pulls - fold the combined-program re-research pulls into the per-program
research-result CSVs that the build already reads.

The weak/duplicate re-research pulls came back as ONE workbook each spanning all three
programs (a Program column + a results sheet). This reads them from research_pulls/,
splits by Program, and UPSERTS each row (keyed by Subawardee UEI, latest pull wins) into

    extracted/<program>_subaward_research_results.csv

so build_program_vendors / build_uei_dimensions pick up the new Role/Description, Source
URLs and NAICS with no other change. Idempotent (re-running replaces, never duplicates).
Raw pulls are read-only.

Run: python3 scripts/merge_research_pulls.py
"""
from __future__ import annotations

import csv
from pathlib import Path

import openpyxl

from _paths import REFACTOR  # noqa: E402
RAW_DIR = REFACTOR / "research_pulls"
EXTRACTED = REFACTOR / "extracted"

LABEL_TO_KEY = {"DDG": "ddg", "Virginia": "virginia", "Columbia": "columbia"}
OUT_COLS = ["Subawardee UEI", "Role / Description", "Source URLs",
            "NAICS-6 code", "NAICS-6 title"]
SIG = {"Program", "Subawardee UEI", "Role / Description"}

# combined-program pulls to fold in, in precedence order (later wins on a UEI clash)
COMBINED_PULLS = [
    "weak_description_completed.xlsx",
    "duplicate_parent_description_completed.xlsx",
]

# single-program pulls (UEI-keyed, no Program column): (raw file, program key). Applied
# after the combined pulls, so a UEI already present is overwritten (latest pass wins).
SINGLE_PROGRAM_PULLS = [
    ("columbia_hedge_secondpass_completed.xlsx", "columbia"),
]


def s(v) -> str:
    return "" if v is None else str(v).strip()


def find_results_sheet(wb):
    for ws in wb.worksheets:
        if SIG <= {s(c.value) for c in ws[1]}:
            return ws
    raise SystemExit(f"no results sheet (headers {SIG}) in {wb}")


def read_combined(fname: str) -> dict[str, list[dict]]:
    """{program_key: [row dicts]} from one combined-program pull."""
    wb = openpyxl.load_workbook(RAW_DIR / fname, read_only=True, data_only=True)
    ws = find_results_sheet(wb)
    idx = {s(c.value): i for i, c in enumerate(ws[1])}
    out: dict[str, list[dict]] = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        uei = s(row[idx["Subawardee UEI"]])
        prog = LABEL_TO_KEY.get(s(row[idx["Program"]]))
        if not uei or not prog:
            continue
        out.setdefault(prog, []).append({c: (s(row[idx[c]]) if c in idx else "") for c in OUT_COLS})
    return out


def read_single_program(fname: str) -> list[dict]:
    """[row dicts] from one single-program pull (UEI-keyed, no Program column)."""
    wb = openpyxl.load_workbook(RAW_DIR / fname, read_only=True, data_only=True)
    sig = {"Subawardee UEI", "Role / Description"}
    ws = next((w for w in wb.worksheets if sig <= {s(c.value) for c in w[1]}), None)
    if ws is None:
        raise SystemExit(f"no results sheet (headers {sig}) in {fname}")
    idx = {s(c.value): i for i, c in enumerate(ws[1])}
    out: list[dict] = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not s(row[idx["Subawardee UEI"]]):
            continue
        out.append({c: (s(row[idx[c]]) if c in idx else "") for c in OUT_COLS})
    return out


def upsert(program: str, new_rows: list[dict]) -> tuple[int, int]:
    path = EXTRACTED / f"{program}_subaward_research_results.csv"
    existing: dict[str, dict] = {}
    if path.exists():
        with path.open(encoding="utf-8-sig", newline="") as fh:
            for r in csv.DictReader(fh):
                existing[r["Subawardee UEI"]] = {c: r.get(c, "") for c in OUT_COLS}
    added = updated = 0
    for r in new_rows:
        uei = r["Subawardee UEI"]
        if uei in existing:
            updated += 1
        else:
            added += 1
        existing[uei] = r
    with path.open("w", encoding="utf-8", newline="") as fh:
        w = csv.DictWriter(fh, fieldnames=OUT_COLS)
        w.writeheader()
        w.writerows(existing.values())
    return added, updated


def main() -> None:
    merged: dict[str, list[dict]] = {}

    def fold(prog: str, rows: list[dict]) -> None:
        merged.setdefault(prog, [])
        seen = {r["Subawardee UEI"] for r in rows}   # later pull wins on a UEI clash
        merged[prog] = [r for r in merged[prog] if r["Subawardee UEI"] not in seen] + rows

    for fname in COMBINED_PULLS:
        for prog, rows in read_combined(fname).items():
            fold(prog, rows)
    for fname, prog in SINGLE_PROGRAM_PULLS:
        fold(prog, read_single_program(fname))
    for prog in ("ddg", "virginia", "columbia"):
        rows = merged.get(prog, [])
        added, updated = upsert(prog, rows)
        print(f"{prog:9}: +{added} new, {updated} replaced  ({len(rows)} rows from pulls)")


if __name__ == "__main__":
    main()
