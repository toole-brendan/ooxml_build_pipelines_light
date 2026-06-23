"""merge_archetype_pulls - fold the completed archetype-classification pulls into a
per-program archetype-result CSV that build_program_vendors reads.

Each program has TWO single-program pulls, one per published axis:
    <program>_capability_domain_completed.xlsx   -> Capability Domain (D) + Basis + URLs
    <program>_primary_output_completed.xlsx      -> Primary Output (P) + Basis + URLs

Both are UEI-keyed within one program. This reads them from research_pulls/, finds the
classified sheet by header signature (the main sheet carries Role / Description, which
disambiguates it from the slimmer '... Output' subset sheet some files also ship), and
merges the two axes per program into

    extracted/<program>_archetype_results.csv
    (Subawardee UEI | Capability Domain (D) | Capability Domain Basis |
     Capability Domain URLs | Primary Output (P) | Primary Output Basis |
     Primary Output URLs)

build_program_vendors then joins on UEI to fill the four archetype columns (the codes +
the 'Research override' basis tier) and the two transient note columns (reasoning +
sources, folded into hover Notes on the Basis cells). The free-text basis and the URLs
are kept in SEPARATE columns here; build_program_vendors composes the note text.

Faithful dump: the per-row Basis prose and URLs are carried verbatim, validation lives
downstream. Idempotent (a re-run overwrites); raw pulls are read-only.

Run: python3 scripts/merge_archetype_pulls.py
"""
from __future__ import annotations

import csv
from pathlib import Path

import openpyxl

from _paths import REFACTOR  # noqa: E402
RAW_DIR = REFACTOR / "research_pulls"
EXTRACTED = REFACTOR / "extracted"

PROGRAMS = ["ddg", "virginia", "columbia"]

# axis key -> (raw-file suffix, code header in the pull, basis header in the pull)
AXES = [
    ("D", "capability_domain", "Capability Domain (D)", "Capability Domain Basis"),
    ("P", "primary_output",    "Primary Output (P)",    "Primary Output Basis"),
]

OUT_COLS = [
    "Subawardee UEI",
    "Capability Domain (D)", "Capability Domain Basis", "Capability Domain URLs",
    "Primary Output (P)",    "Primary Output Basis",    "Primary Output URLs",
]


def s(v) -> str:
    return "" if v is None else str(v).strip()


def find_classified_sheet(wb, code_header: str, basis_header: str):
    """The main classified sheet by header signature: it carries the axis code +
    basis columns AND Role / Description (which the slim '... Output' subset sheet
    lacks), so the signature can't accidentally pick the subset sheet."""
    sig = {"Subawardee UEI", "Role / Description", code_header, basis_header}
    for ws in wb.worksheets:
        if sig <= {s(c.value) for c in ws[1]}:
            return ws
    raise SystemExit(f"no classified sheet (needs {sig}) in {wb}")


def url_col(headers: list[str]) -> int:
    """Index of the source-URL column; its header name varies across pulls
    ('URL Source(s)' / 'Source URL(s)' / 'URL source(s)'), so match on 'url'."""
    for i, h in enumerate(headers):
        if "url" in s(h).lower():
            return i
    raise SystemExit(f"no URL column in headers {headers}")


def read_axis(path: Path, code_header: str, basis_header: str) -> dict[str, tuple[str, str, str]]:
    """uei -> (code, basis_text, urls) for one axis pull."""
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = find_classified_sheet(wb, code_header, basis_header)
    rows = list(ws.iter_rows(values_only=True))
    headers = [s(c) for c in rows[0]]
    ui = headers.index("Subawardee UEI")
    ci = headers.index(code_header)
    bi = headers.index(basis_header)
    li = url_col(headers)
    out: dict[str, tuple[str, str, str]] = {}
    for r in rows[1:]:
        if not any(c is not None for c in r):
            continue
        uei = s(r[ui])
        if uei:
            out[uei] = (s(r[ci]), s(r[bi]), s(r[li]))
    wb.close()
    return out


def merge_program(program: str) -> None:
    by_uei: dict[str, dict[str, str]] = {}
    counts = {}
    for axis, suffix, code_header, basis_header in AXES:
        path = RAW_DIR / f"{program}_{suffix}_completed.xlsx"
        if not path.exists():
            print(f"  [{program}] {axis}: MISSING {path.name} - skipped")
            counts[axis] = 0
            continue
        rows = read_axis(path, code_header, basis_header)
        counts[axis] = len(rows)
        for uei, (code, basis, urls) in rows.items():
            rec = by_uei.setdefault(uei, {c: "" for c in OUT_COLS})
            rec["Subawardee UEI"] = uei
            if axis == "D":
                rec["Capability Domain (D)"] = code
                rec["Capability Domain Basis"] = basis
                rec["Capability Domain URLs"] = urls
            else:
                rec["Primary Output (P)"] = code
                rec["Primary Output Basis"] = basis
                rec["Primary Output URLs"] = urls

    out_path = EXTRACTED / f"{program}_archetype_results.csv"
    with out_path.open("w", encoding="utf-8", newline="") as fh:
        w = csv.DictWriter(fh, fieldnames=OUT_COLS)
        w.writeheader()
        for uei in sorted(by_uei):
            w.writerow(by_uei[uei])

    d_only = sum(1 for r in by_uei.values()
                 if r["Capability Domain (D)"] and not r["Primary Output (P)"])
    p_only = sum(1 for r in by_uei.values()
                 if r["Primary Output (P)"] and not r["Capability Domain (D)"])
    print(f"==== {program.upper()} archetype merge ====")
    print(f"output: {out_path}")
    print(f"  Capability Domain rows : {counts.get('D', 0)}")
    print(f"  Primary Output rows    : {counts.get('P', 0)}")
    print(f"  distinct UEIs (union)  : {len(by_uei)}  (D-only {d_only} / P-only {p_only})")


if __name__ == "__main__":
    for prog in PROGRAMS:
        merge_program(prog)
