"""tools/style_audit.py - house-style linter for the BUILT workbook.

NOT part of the build pipeline (the pipeline is strictly stdlib-only raw OOXML).
Like validate_workbook.py, this reads back the ALREADY-BUILT xlsx with openpyxl and
enforces the workbook_core house style the readability audit called for. Run it on a
FRESH source build (not an Excel-resaved copy) so the formula checks see the real
formula strings.

Usage:
    python tools/style_audit.py        # lint; exit 1 on any hard failure
Honors WB_OUT (same as build_workbook.py / validate_workbook.py).

Checks
  Hard failures (exit 1):
    - B2 title equals the tab name.
    - Every gutter-'x' section banner matches  ^§\\d+[a-z]? - .+
    - A gutter 'x' exists only where outlined detail follows it.
    - Every outlined section sits under a gutter-'x' banner.
    - No en/em dashes survive in any cell literal or formula.
    - No forbidden formula-helper header is left visible (must be a hidden column).
    - No invalid cross-section reference (e.g. 'Taxonomy §5' when Taxonomy has §1-§4).
  Warnings (reported, exit unaffected):
    - Row-3 caption exceeds 120 characters.
    - Section title exceeds 60 characters.
    - A visible column is wider than 44 (outside the allowlist).
    - Visible text begins with spaces.
    - Visible prose carries an implementation term (-> , <-- , => , .csv, INDEX, ...).
"""
from __future__ import annotations

import os
import re
import sys
from pathlib import Path

import openpyxl

# Built workbook lands at the project root, two levels up from tools/.
OUT = Path(__file__).resolve().parents[2] / os.environ.get(
    "WB_OUT", "award_classification_refactor.xlsx")

SECTION_RE = re.compile(r"^§\d+[a-z]? - .+")
DASH_RE = re.compile(r"[–—]")          # en dash, em dash
CROSSREF_RE = re.compile(r"\b(Taxonomy|Methodology) §(\d+)")
# Highest real section number on the referenced tabs (drives the cross-ref check).
SECTION_MAX = {"Taxonomy": 4, "Methodology": 6}

# Helper columns that must never be visible (kept in-grid for formulas, hidden).
FORBIDDEN_HEADERS = {
    "SM Match Row", "Override Match Row", "NAICS Map Match Row", "SWBS Match Row", "Key",
}

# Implementation / pipeline language that should not appear in visible prose.
IMPL_TERMS = [
    "reviewer finding", "The build", ".csv", "extracted/", "scope-load",
    "include=Y", "include=N", "MATCH", "INDEX", "SUMIFS", "helper",
    "REMOVED", "placeholder", "->", "<--", "=>",
]

CAPTION_MAX = 120
SECTION_TITLE_MAX = 60
WIDTH_MAX = 44
# Columns intentionally wider than WIDTH_MAX (deliberate prose columns).
WIDTH_ALLOW = {("Taxonomy", "D")}      # the code-definition column


def _is_formula(v) -> bool:
    return isinstance(v, str) and v.startswith("=")


def _text_cells(ws):
    """Yield (cell, text) for every non-empty, non-formula string cell (skip the 'x')."""
    for row in ws.iter_rows():
        for cell in row:
            v = cell.value
            if isinstance(v, str) and v and not _is_formula(v) and v != "x":
                yield cell, v


def audit() -> int:
    wb = openpyxl.load_workbook(OUT, data_only=False)
    fails: list[str] = []
    warns: list[str] = []

    for ws in wb.worksheets:
        tab = ws.title

        # --- B2 title equals the tab name ----------------------------------------
        b2 = ws["B2"].value
        if b2 != tab:
            fails.append(f"[{tab}] B2 title {b2!r} != tab name {tab!r}")

        # --- Row-3 caption length (warn) -----------------------------------------
        cap = ws["B3"].value
        if isinstance(cap, str) and len(cap) > CAPTION_MAX:
            warns.append(f"[{tab}] row-3 caption {len(cap)} chars (>{CAPTION_MAX})")

        # --- Section banners (gutter 'x' rows) + outline coupling ----------------
        x_rows = [c.row for c in ws["A"] if c.value == "x"]
        outlined = {r: (ws.row_dimensions[r].outline_level or 0) for r in range(1, ws.max_row + 1)}
        outlined_rows = sorted(r for r, lvl in outlined.items() if lvl and lvl > 0)

        for i, r in enumerate(x_rows):
            text = ws.cell(row=r, column=2).value
            if not isinstance(text, str) or not SECTION_RE.match(text):
                fails.append(f"[{tab}] section banner at row {r} not '§N - ...': {text!r}")
            elif len(text) > SECTION_TITLE_MAX:
                warns.append(f"[{tab}] section title {len(text)} chars (>{SECTION_TITLE_MAX}): {text!r}")
            # outlined detail must follow this banner before the next one
            nxt = x_rows[i + 1] if i + 1 < len(x_rows) else ws.max_row + 1
            if not any(r < o < nxt for o in outlined_rows):
                fails.append(f"[{tab}] gutter 'x' at row {r} has no outlined detail below it")

        # every outlined row must sit under some gutter-'x' banner above it
        for o in outlined_rows:
            if not any(x < o for x in x_rows):
                fails.append(f"[{tab}] outlined row {o} has no gutter-'x' section banner above it")
                break

        # --- Dashes in any literal / formula (hard) ------------------------------
        for row in ws.iter_rows():
            for cell in row:
                if isinstance(cell.value, str) and DASH_RE.search(cell.value):
                    kind = "formula" if _is_formula(cell.value) else "literal"
                    fails.append(f"[{tab}] en/em dash in {kind} at {cell.coordinate}: {cell.value[:60]!r}")

        # --- Forbidden helper headers must be hidden -----------------------------
        for cell, v in _text_cells(ws):
            if v in FORBIDDEN_HEADERS:
                col = cell.column_letter
                if not (ws.column_dimensions[col].hidden):
                    fails.append(f"[{tab}] forbidden helper header {v!r} visible at {cell.coordinate}")

        # --- Cross-section references (hard) -------------------------------------
        for cell, v in _text_cells(ws):
            for sheet_ref, num in CROSSREF_RE.findall(v):
                if int(num) > SECTION_MAX.get(sheet_ref, 99):
                    fails.append(f"[{tab}] invalid cross-ref {sheet_ref} §{num} at {cell.coordinate}")

        # --- Implementation terms + leading spaces (warn) ------------------------
        for cell, v in _text_cells(ws):
            # a pure-whitespace cell is the table's right-spacer (a clipping device),
            # not prose - only flag text that has content AND a leading space.
            if v[:1] == " " and v.strip():
                warns.append(f"[{tab}] visible text begins with space at {cell.coordinate}: {v[:40]!r}")
            for term in IMPL_TERMS:
                if term in v:
                    warns.append(f"[{tab}] implementation term {term!r} at {cell.coordinate}: {v[:50]!r}")

        # --- Column widths (warn) ------------------------------------------------
        for col, dim in ws.column_dimensions.items():
            if dim.width and dim.width > WIDTH_MAX and not dim.hidden \
                    and (tab, col) not in WIDTH_ALLOW:
                warns.append(f"[{tab}] column {col} width {dim.width:g} (>{WIDTH_MAX})")

    print(f"file: {OUT.name}")
    print(f"hard failures: {len(fails)}")
    for f in fails:
        print("   FAIL ", f)
    print(f"warnings: {len(warns)}")
    for w in warns:
        print("   warn ", w)
    return 1 if fails else 0


if __name__ == "__main__":
    sys.exit(audit())
