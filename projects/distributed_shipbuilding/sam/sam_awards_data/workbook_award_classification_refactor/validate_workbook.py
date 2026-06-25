"""QA / validation tool for the built workbook.

NOT part of the build pipeline. The pipeline (workbook_award_classification_refactor/) is
strictly stdlib-only raw OOXML; this script uses openpyxl purely to read back the
*already-built* file and sanity-check it. Never import this from a sheet
module or from lib.py.

Usage:
    python validate_workbook.py                       # structural report
    python validate_workbook.py "Taxonomy!B2" ...     # dump cells

Honors WB_OUT (same as the build) so it validates whichever file the build wrote.

openpyxl does not evaluate formulas (only Excel/LibreOffice do), so cell
dumps show the formula string for formula cells and the literal for
hardcoded cells - which is exactly what we want to audit the build.
"""
from __future__ import annotations

import os
import sys
import xml.dom.minidom as minidom
import zipfile
from pathlib import Path

# Built workbook lands at the project root (projects/distributed_shipbuilding/sam/sam_awards_data/), one level up from here.
OUT = Path(__file__).resolve().parent.parent / os.environ.get(
    "WB_OUT", "20260620_Distributed Shipbuilding Master SAM_vS.xlsx")

# Expected tab order (group-contiguous). A drift here means a sheet was added/removed/reordered.
EXPECTED_SHEETS = [
    "Executive Summary",
    "Domain Concentration",
    "Where to Play",
    "Subaward Activity",
    "Market Bridge",
    "Taxonomy",
    "Methodology",

    "Supplier Master",
    "Supplier-Year Activity",
    "DDG Program Vendors",
    "DDG SWBS by Ship-System",
    "Virginia Program Vendors",
    "Columbia Program Vendors",

    "Mapping - NAICS Defaults",
    "Mapping - Vendor Overrides",
    "Mapping - HII Code to SWBS",
    "Deflators",

    "Prime Awards",
    "DDG Subaward Transactions",
    "Virginia Subaward Transactions",
    "Columbia Subaward Transactions",
]

# Muted per-group tab palette (lib.py overrides workbook_core.groups). Anchors, one per group.
EXPECTED_TAB_COLORS = {
    "Executive Summary": "262626", "Where to Play": "262626",      # summary -> charcoal
    "Domain Concentration": "262626",
    "Taxonomy": "2C5E5E", "Methodology": "2C5E5E",                 # guide -> muted teal
    "Deflators": "556B2F",                                         # inputs -> olive
    "Supplier Master": "48596B", "Supplier-Year Activity": "48596B",  # model -> slate
    "Prime Awards": "203864",                                     # data -> navy
}

# (sheet, a visible formula-column header) spot checks for the new analytical sheets: the cell
# under the header on the first data row (row 9) must carry a live formula.
EXPECTED_FORMULA_COLS = [
    ("Supplier-Year Activity", "Activity Status"),
    ("Where to Play", "Structure Class"),
]


def _tab_rgb(ws) -> str:
    tc = ws.sheet_properties.tabColor
    if tc is None:
        return ""
    rgb = getattr(tc, "rgb", None) or (tc if isinstance(tc, str) else "")
    return (rgb or "")[-6:].upper()


def _spot_formula_ok(ws, header: str) -> bool:
    """The cell on row 9 under `header` (found in the row-8 header row) is a formula."""
    for cell in ws[8]:
        if cell.value == header:
            v = ws.cell(row=9, column=cell.column).value
            return isinstance(v, str) and v.startswith("=")
    return False


def structural_report() -> int:
    z = zipfile.ZipFile(OUT)
    bad = 0
    for n in z.namelist():
        if n.endswith((".xml", ".rels")):
            try:
                minidom.parseString(z.read(n))
            except Exception as e:  # noqa: BLE001
                print(f"  XML ERROR  {n}: {e}")
                bad += 1
    print(f"file: {OUT.name}  ({OUT.stat().st_size:,} bytes)")
    print(f"parts: {len(z.namelist())}  | xml errors: {bad}")

    import openpyxl

    wb = openpyxl.load_workbook(OUT, data_only=False)
    print(f"sheets ({len(wb.sheetnames)}): {wb.sheetnames}")
    # Scan every cell for error literals that survived the build.
    errors: list[str] = []
    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for c in row:
                if isinstance(c.value, str) and c.value.startswith("#") and c.value.endswith("!"):
                    errors.append(f"{ws.title}!{c.coordinate} = {c.value}")
    print(f"error-literal cells: {len(errors)}")
    for e in errors[:20]:
        print("   ", e)

    # --- expected sheet set/order ---------------------------------------------
    checks: list[str] = []
    if wb.sheetnames != EXPECTED_SHEETS:
        miss = [s for s in EXPECTED_SHEETS if s not in wb.sheetnames]
        extra = [s for s in wb.sheetnames if s not in EXPECTED_SHEETS]
        checks.append(f"sheet set/order drift: missing={miss} extra={extra}")

    # --- tab palette ----------------------------------------------------------
    for tab, want in EXPECTED_TAB_COLORS.items():
        if tab not in wb.sheetnames:
            checks.append(f"tab-color: missing sheet {tab!r}")
            continue
        got = _tab_rgb(wb[tab])
        if got != want:
            checks.append(f"tab-color: {tab} is {got!r} != {want}")

    # --- new analytical sheets carry live formulas ----------------------------
    for tab, header in EXPECTED_FORMULA_COLS:
        if tab not in wb.sheetnames:
            checks.append(f"formula-col: missing sheet {tab!r}")
        elif not _spot_formula_ok(wb[tab], header):
            checks.append(f"formula-col: {tab!r} has no live formula under {header!r}")

    print(f"structural checks: {len(checks)} issue(s)")
    for ck in checks:
        print("   ", ck)
    return 1 if (bad or errors or checks) else 0


def dump_cells(refs: list[str]) -> int:
    import openpyxl

    wb = openpyxl.load_workbook(OUT, data_only=False)
    for ref in refs:
        sheet, _, cell = ref.rpartition("!")
        sheet = sheet.strip("'")
        if sheet not in wb.sheetnames:
            print(f"{ref:<28} -> NO SUCH SHEET (have {wb.sheetnames})")
            continue
        v = wb[sheet][cell].value
        print(f"{ref:<28} -> {v!r}")
    return 0


if __name__ == "__main__":
    if len(sys.argv) > 1:
        sys.exit(dump_cells(sys.argv[1:]))
    sys.exit(structural_report())
