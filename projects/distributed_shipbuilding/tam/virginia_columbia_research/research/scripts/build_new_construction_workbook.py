#!/usr/bin/env python3
"""
Build the submarine new-construction subaward workbook.

Inputs (must exist):
  extracted/nc_annual_by_piid.csv
  extracted/nc_annual_by_vendor.csv
  extracted/nc_lifetime_vendors.csv
  extracted/nc_records_long.csv
  extracted/nc_scope_summary.json
  extracted/entity_naics_lookup.csv      <-- from pull_sam_entity_naics.py
  extracted/scn_p5c_per_fy_reconciled.csv

Output:
  extracted/submarine_new_construction_subawards.xlsx
"""
import csv
import json
from collections import defaultdict
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

REPO = Path("/Users/brendantoole/projects2/submarine_outsourced_work")
EXT = REPO / "extracted"
OUT_XLSX = EXT / "submarine_new_construction_subawards.xlsx"

# ---- Style helpers ---------------------------------------------------------

BOLD = Font(bold=True)
TITLE = Font(bold=True, size=14)
HEADER_FILL = PatternFill(start_color="1F3864", end_color="1F3864", fill_type="solid")
HEADER_FONT = Font(bold=True, color="FFFFFF")
SUBHEADER_FILL = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
WARN_FILL = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
THIN = Side(style="thin", color="BFBFBF")
BOX = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def header_row(ws, row, values, fill=HEADER_FILL, font=HEADER_FONT):
    for c, v in enumerate(values, start=1):
        cell = ws.cell(row=row, column=c, value=v)
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = BOX


def title_row(ws, row, col, text):
    cell = ws.cell(row=row, column=col, value=text)
    cell.font = TITLE


def auto_width(ws, widths):
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w


# ---- Data loaders ----------------------------------------------------------

def load_scope():
    return json.load(open(EXT / "nc_scope_summary.json"))


def load_by_piid():
    with open(EXT / "nc_annual_by_piid.csv") as f:
        return list(csv.DictReader(f))


def load_by_vendor():
    with open(EXT / "nc_annual_by_vendor.csv") as f:
        return list(csv.DictReader(f))


def load_lifetime():
    with open(EXT / "nc_lifetime_vendors.csv") as f:
        return list(csv.DictReader(f))


def load_naics():
    path = EXT / "entity_naics_lookup.csv"
    if not path.exists():
        return []
    with open(path) as f:
        return list(csv.DictReader(f))


def load_trends():
    path = EXT / "nc_trends.csv"
    if not path.exists():
        return []
    return list(csv.DictReader(open(path)))


def load_geo_state():
    path = EXT / "nc_geo_by_state.csv"
    if not path.exists():
        return []
    return list(csv.DictReader(open(path)))


def load_geo_country():
    path = EXT / "nc_geo_by_country.csv"
    if not path.exists():
        return []
    return list(csv.DictReader(open(path)))


def load_foreign_share_by_fy():
    path = EXT / "nc_foreign_share_by_fy.csv"
    if not path.exists():
        return []
    return list(csv.DictReader(open(path)))


def load_hii_context():
    path = EXT / "hii_context.csv"
    if not path.exists():
        return []
    return list(csv.DictReader(open(path)))


def load_curated_quotes():
    path = EXT / "hii_curated_quotes.csv"
    if not path.exists():
        return []
    return list(csv.DictReader(open(path)))


def load_basic_construction():
    """Return {FY (int): {'Va': $M, 'Col': $M, 'Total': $M}}."""
    out = defaultdict(lambda: {"Va": 0.0, "Col": 0.0, "Total": 0.0})
    with open(EXT / "scn_p5c_per_fy_reconciled.csv") as f:
        for r in csv.DictReader(f):
            if r["Cost Category"] != "Basic Construction/Conversion":
                continue
            fy = int(r["FY"])
            amt = float(r["Best Value $M"])
            if r["LI"] == "2013":
                out[fy]["Va"] += amt
            elif r["LI"] == "1045":
                out[fy]["Col"] += amt
            out[fy]["Total"] = out[fy]["Va"] + out[fy]["Col"]
    return dict(out)


# ---- Sheet builders --------------------------------------------------------

def sheet_cover(wb, scope):
    ws = wb.create_sheet("Cover", 0)
    pull_date = datetime.now().strftime("%Y-%m-%d")
    rows = [
        ("Submarine New-Construction Subawards — FY-by-FY + Work-Type Breakdown", "title"),
        ("", ""),
        ("Purpose", "section"),
        ("Quantify first-tier subaward dollars per fiscal year for U.S. submarine NEW-CONSTRUCTION prime contracts, broken down by work type (NAICS), recipient vendor, and prime PIID.", ""),
        ("", ""),
        ("Scope", "section"),
        (f"In-scope PIIDs: {len(scope['in_scope_piids'])} new-construction prime contracts (GDEB construction masters + design/concept + VPM components + BPMI reactor GFE + LM/BAE/RR other GFE).", ""),
        (f"Out-of-scope PIIDs excluded: 2 (Hartford EOH overhaul + Va Tech Instructions/HPAD backfit).", ""),
        (f"Recipient exclusions: 3 MIB / workforce pass-through entities (BlueForge Alliance, Training Modernization Group, Institute for Advanced Learning and Research).", ""),
        (f"Time window: FY{scope['fy_range'][0]} – FY{scope['fy_range'][1]} (action-date-based).", ""),
        ("", ""),
        ("Data sources", "section"),
        ("Subaward records: SAM.gov Acquisition Subaward Reporting API (/contract/v1/subcontracts/search, status=Published).", ""),
        ("NAICS enrichment: SAM.gov Entity Management API (/entity-information/v3/entities, primary NAICS).", ""),
        ("Basic Construction denominator: SCN P-5c cost-category reconciliation across PB22-PB27 books (uses most-revised vintage per FY).", ""),
        ("", ""),
        ("Headline numbers", "section"),
        (f"Records kept (in-scope, MIB stripped): {scope['records_kept']:,}", ""),
        (f"Dollars kept (in-scope, MIB stripped): ${scope['total_dollars_in_scope_$M']:,.1f}M", ""),
        (f"Dollars excluded as MIB pass-through:  ${scope['dollars_excluded_mib_$M']:,.1f}M (separately tracked)", ""),
        (f"Unique parent UEIs in scope: {scope['unique_parent_ueis_in_scope']:,}", ""),
        ("", ""),
        ("Pull date", "section"),
        (pull_date, ""),
        ("", ""),
        ("Sheets in this workbook", "section"),
        ("Scope_PIIDs       — the 15 in-scope PIIDs (+ 2 excluded with reason)", ""),
        ("FY_Headline       — primary answer: $ subbed per FY vs SCN Basic Construction allocated, ratio %", ""),
        ("FY_By_NAICS       — work-type breakdown via vendor primary NAICS (4-digit industry group)", ""),
        ("FY_By_Vendor      — top vendors per FY (parent-UEI rolled, MIB stripped)", ""),
        ("FY_By_PIID        — which prime contract is driving the cash flow each FY", ""),
        ("Lifetime_Vendors  — top 100 vendors across the full window (with NAICS where enriched)", ""),
        ("NAICS_Lookup      — the UEI → primary NAICS map (audit trail)", ""),
        ("Caveats           — known limitations and methodology notes", ""),
        ("", ""),
        ("Key caveats (see Caveats tab for detail)", "section"),
        ("1. FFATA reporting lag depresses FY25 (~50-70% complete) and FY26 (~0-5% complete).", "warn"),
        ("2. HII-NNS team-build share (~50% of Va, ~22% of Col) is largely invisible — flows through GDEB as vendor of record, not as FFATA subs.", "warn"),
        ("3. NAICS is the vendor's primary self-reported classifier — reflects what they ARE, not necessarily what THIS sub was for. Top-150 vendors enriched (~90% of $); long tail bucketed as 'Unenriched'.", "warn"),
        ("4. Basic Construction denominator follows per-ship procurement cadence — Columbia has no Va-class-like annual flow; FY21/24/26/27 are Col procurement years.", "warn"),
    ]
    for i, (text, kind) in enumerate(rows, start=1):
        cell = ws.cell(row=i, column=1, value=text)
        if kind == "title":
            cell.font = Font(bold=True, size=16)
        elif kind == "section":
            cell.font = Font(bold=True, size=12, color="1F3864")
        elif kind == "warn":
            cell.fill = WARN_FILL
        cell.alignment = Alignment(wrap_text=True, vertical="top")
    ws.column_dimensions["A"].width = 140


def sheet_scope_piids(wb, scope):
    ws = wb.create_sheet("Scope_PIIDs")
    title_row(ws, 1, 1, "In-scope new-construction PIIDs (15)")
    header_row(ws, 3, ["PIID", "Prime", "Class", "Label"])
    row = 4
    for piid, meta in sorted(scope["in_scope_piids"].items()):
        ws.cell(row=row, column=1, value=piid).font = Font(name="Menlo", size=11)
        ws.cell(row=row, column=2, value=meta["prime"])
        ws.cell(row=row, column=3, value=meta["class"])
        ws.cell(row=row, column=4, value=meta["label"])
        row += 1

    row += 2
    title_row(ws, row, 1, "Excluded PIIDs (in our pull, but not new construction)")
    row += 2
    header_row(ws, row, ["PIID", "Prime", "Reason for exclusion"])
    row += 1
    for piid, meta in sorted(scope["out_of_scope_piids"].items()):
        ws.cell(row=row, column=1, value=piid).font = Font(name="Menlo", size=11)
        ws.cell(row=row, column=2, value=meta["prime"])
        ws.cell(row=row, column=3, value=meta["reason"])
        ws.cell(row=row, column=3).fill = WARN_FILL
        row += 1

    row += 2
    title_row(ws, row, 1, "MIB / workforce pass-through vendors excluded (recipient side)")
    row += 2
    header_row(ws, row, ["UEI", "Vendor name", "Notes"])
    row += 1
    notes = {
        "F8PEZKXES8B1": "MIB consortium — workforce / supplier infrastructure distribution",
        "QLJZVM6XKR71": "Workforce training",
        "TCM3R4JPRKY4": "Workforce / STEM education",
    }
    for uei, name in scope["excluded_mib_ueis"].items():
        ws.cell(row=row, column=1, value=uei).font = Font(name="Menlo", size=11)
        ws.cell(row=row, column=2, value=name)
        ws.cell(row=row, column=3, value=notes.get(uei, ""))
        ws.cell(row=row, column=3).fill = WARN_FILL
        row += 1

    auto_width(ws, [22, 10, 16, 80])


def sheet_fy_headline(wb, by_piid, basic_construction):
    """The primary answer table: $ subbed per FY × Basic Construction → ratio."""
    ws = wb.create_sheet("FY_Headline")
    title_row(ws, 1, 1, "FY headline — new-construction subawards $ vs SCN Basic Construction allocated")
    ws.cell(row=2, column=1, value="(MIB pass-throughs already excluded. FY by subAwardDate action FY.)").font = Font(italic=True, color="666666")

    header_row(ws, 4, [
        "FY", "Va BCC $M", "Col BCC $M", "Total BCC $M",
        "Visible subs $M", "% of BCC", "# records", "Notes",
    ])
    row = 5
    for r in by_piid:
        fy = int(r["FY"])
        subs = float(r["FY_TOTAL_$M"])
        bcc = basic_construction.get(fy, {"Va": 0.0, "Col": 0.0, "Total": 0.0})
        ratio = subs / bcc["Total"] if bcc["Total"] > 0 else None
        notes = []
        if bcc["Col"] == 0 and fy >= 2021:
            notes.append("no Col procurement this FY")
        if fy >= 2025:
            notes.append("FFATA lag — figures will rise")

        ws.cell(row=row, column=1, value=fy)
        ws.cell(row=row, column=2, value=bcc["Va"] if bcc["Va"] else None)
        ws.cell(row=row, column=3, value=bcc["Col"] if bcc["Col"] else None)
        ws.cell(row=row, column=4, value=bcc["Total"] if bcc["Total"] else None)
        ws.cell(row=row, column=5, value=subs)
        ws.cell(row=row, column=6, value=ratio)
        ws.cell(row=row, column=7, value=int(r["FY_TOTAL_count"]))
        ws.cell(row=row, column=8, value="; ".join(notes) if notes else "")

        for c in (2, 3, 4, 5):
            ws.cell(row=row, column=c).number_format = "#,##0.0"
        ws.cell(row=row, column=6).number_format = "0.0%"
        if "lag" in (ws.cell(row=row, column=8).value or ""):
            for c in range(1, 9):
                ws.cell(row=row, column=c).fill = WARN_FILL
        row += 1

    # Cumulative FY22-FY24 (most stable window) summary block
    row += 2
    title_row(ws, row, 1, "Cumulative window summary (most-stable FYs)")
    row += 2
    header_row(ws, row, ["Window", "Sum BCC $M", "Sum subs $M", "Subs / BCC %", "Notes"])
    row += 1
    for window_label, fys in [("FY22-FY24 (stable)", [2022, 2023, 2024]),
                                ("FY22-FY26 (full available)", [2022, 2023, 2024, 2025, 2026])]:
        bcc_sum = sum(basic_construction.get(y, {"Total": 0.0})["Total"] for y in fys)
        subs_sum = sum(float(r["FY_TOTAL_$M"]) for r in by_piid if int(r["FY"]) in fys)
        ratio = subs_sum / bcc_sum if bcc_sum else None
        ws.cell(row=row, column=1, value=window_label)
        ws.cell(row=row, column=2, value=bcc_sum).number_format = "#,##0.0"
        ws.cell(row=row, column=3, value=subs_sum).number_format = "#,##0.0"
        ws.cell(row=row, column=4, value=ratio).number_format = "0.0%"
        ws.cell(row=row, column=5, value=("includes FY25-26 lag" if 2025 in fys else "stable data"))
        row += 1

    auto_width(ws, [6, 12, 12, 12, 14, 10, 11, 40])


def sheet_fy_by_naics(wb, by_vendor, naics_lookup):
    """FY × NAICS 4-digit work-type breakdown."""
    ws = wb.create_sheet("FY_By_NAICS")
    title_row(ws, 1, 1, "FY × work-type (NAICS 4-digit) breakdown")
    ws.cell(row=2, column=1,
            value="Vendor primary NAICS via SAM Entity Management API. "
                  "Unenriched vendors (long-tail by $) bucketed as 'UNENR'. "
                  "MIB pass-throughs already excluded."
            ).font = Font(italic=True, color="666666")

    # Build UEI → (naics_4, naics_desc) map
    uei_to_naics = {}
    for r in naics_lookup:
        n4 = r.get("naics_4digit", "")
        desc = r.get("naics_desc", "")
        if r.get("lookup_status") == "ok" and n4:
            uei_to_naics[r["uei"]] = (n4, desc, r.get("naics_sector_label", ""))

    # Aggregate by FY × NAICS-4
    fy_naics = defaultdict(lambda: defaultdict(float))
    naics_label = {}
    naics_sector = {}
    fys_seen = set()
    for r in by_vendor:
        fy = int(r["fy"])
        fys_seen.add(fy)
        amt = float(r["amount_M"])
        n4, desc, sector = uei_to_naics.get(r["uei"], ("UNENR", "Unenriched (long-tail vendor)", "Unknown"))
        fy_naics[fy][n4] += amt
        if n4 not in naics_label:
            naics_label[n4] = desc
            naics_sector[n4] = sector

    fys_sorted = sorted(fys_seen)
    naics_keys_sorted = sorted(
        naics_label.keys(),
        key=lambda k: (-sum(fy_naics[fy][k] for fy in fys_sorted),)
    )

    # Wide table: rows = NAICS, columns = FYs
    header_row(ws, 4, ["NAICS 4-digit", "Industry", "Sector"] + [f"FY{y}" for y in fys_sorted] + ["Total $M", "% of total"])
    row = 5
    grand_total = sum(sum(fy_naics[fy].values()) for fy in fys_sorted)
    for k in naics_keys_sorted:
        total = sum(fy_naics[fy][k] for fy in fys_sorted)
        if total < 0.01:
            continue
        ws.cell(row=row, column=1, value=k).font = Font(name="Menlo", size=11)
        ws.cell(row=row, column=2, value=naics_label[k][:80])
        ws.cell(row=row, column=3, value=naics_sector[k])
        for i, fy in enumerate(fys_sorted, start=4):
            v = fy_naics[fy][k]
            cell = ws.cell(row=row, column=i, value=v if v else None)
            cell.number_format = "#,##0.0"
        col_total = 4 + len(fys_sorted)
        ws.cell(row=row, column=col_total, value=total).number_format = "#,##0.0"
        ws.cell(row=row, column=col_total + 1, value=total / grand_total if grand_total else 0).number_format = "0.0%"
        if k == "UNENR":
            for c in range(1, col_total + 2):
                ws.cell(row=row, column=c).fill = WARN_FILL
        row += 1

    # Totals row
    ws.cell(row=row, column=1, value="TOTAL").font = BOLD
    for i, fy in enumerate(fys_sorted, start=4):
        ws.cell(row=row, column=i, value=sum(fy_naics[fy].values())).number_format = "#,##0.0"
        ws.cell(row=row, column=i).font = BOLD
    col_total = 4 + len(fys_sorted)
    ws.cell(row=row, column=col_total, value=grand_total).font = BOLD
    ws.cell(row=row, column=col_total).number_format = "#,##0.0"

    auto_width(ws, [10, 50, 30] + [10] * len(fys_sorted) + [12, 11])


def sheet_fy_by_vendor(wb, by_vendor, naics_lookup):
    """Per-FY top vendors."""
    ws = wb.create_sheet("FY_By_Vendor")
    title_row(ws, 1, 1, "Top vendors per FY (parent-UEI rolled, MIB stripped)")

    uei_to_naics = {}
    for r in naics_lookup:
        if r.get("lookup_status") == "ok" and r.get("primary_naics"):
            uei_to_naics[r["uei"]] = (r["primary_naics"], r.get("naics_desc", ""))

    fys = sorted({int(r["fy"]) for r in by_vendor})

    row = 3
    for fy in fys:
        title_row(ws, row, 1, f"FY{fy}")
        row += 1
        header_row(ws, row, ["Rank", "Vendor", "Parent UEI", "Foreign?",
                              "NAICS", "Industry", "$M", "% of FY"])
        row += 1
        fy_rows = sorted(
            [r for r in by_vendor if int(r["fy"]) == fy],
            key=lambda x: -float(x["amount_M"])
        )
        fy_total = sum(float(r["amount_M"]) for r in fy_rows)
        for i, r in enumerate(fy_rows[:20], start=1):
            naics_code, naics_desc = uei_to_naics.get(r["uei"], ("", ""))
            ws.cell(row=row, column=1, value=i)
            ws.cell(row=row, column=2, value=r["vendor"][:60])
            ws.cell(row=row, column=3, value=r["uei"]).font = Font(name="Menlo", size=10)
            ws.cell(row=row, column=4, value=r["foreign"])
            ws.cell(row=row, column=5, value=naics_code).font = Font(name="Menlo", size=10)
            ws.cell(row=row, column=6, value=naics_desc[:50])
            amt = float(r["amount_M"])
            ws.cell(row=row, column=7, value=amt).number_format = "#,##0.00"
            ws.cell(row=row, column=8, value=amt/fy_total if fy_total else 0).number_format = "0.0%"
            row += 1
        # Long tail
        tail_amt = sum(float(r["amount_M"]) for r in fy_rows[20:])
        if tail_amt > 0:
            ws.cell(row=row, column=2, value=f"Long tail (ranks 21+, {len(fy_rows)-20} vendors)").font = Font(italic=True)
            ws.cell(row=row, column=7, value=tail_amt).number_format = "#,##0.00"
            ws.cell(row=row, column=8, value=tail_amt/fy_total if fy_total else 0).number_format = "0.0%"
            row += 1
        # FY total
        ws.cell(row=row, column=2, value="FY TOTAL").font = BOLD
        ws.cell(row=row, column=7, value=fy_total).font = BOLD
        ws.cell(row=row, column=7).number_format = "#,##0.00"
        row += 2

    auto_width(ws, [6, 50, 16, 10, 10, 40, 12, 10])


def sheet_fy_by_piid(wb, by_piid, scope):
    """Per FY × PIID, with class column."""
    ws = wb.create_sheet("FY_By_PIID")
    title_row(ws, 1, 1, "FY × PIID — which prime is driving cash flow each FY")
    ws.cell(row=2, column=1, value="Long-form table; one row per FY × PIID. Zero rows omitted.").font = Font(italic=True, color="666666")

    header_row(ws, 4, ["FY", "PIID", "Prime", "Class", "Label", "$M", "# records"])
    row = 5
    for r in by_piid:
        fy = int(r["FY"])
        for piid in sorted(scope["in_scope_piids"].keys()):
            key = f"{piid}_$M"
            cnt_key = f"{piid}_count"
            amt = float(r.get(key, 0) or 0)
            cnt = int(r.get(cnt_key, 0) or 0)
            if amt == 0 and cnt == 0:
                continue
            meta = scope["in_scope_piids"][piid]
            ws.cell(row=row, column=1, value=fy)
            ws.cell(row=row, column=2, value=piid).font = Font(name="Menlo", size=11)
            ws.cell(row=row, column=3, value=meta["prime"])
            ws.cell(row=row, column=4, value=meta["class"])
            ws.cell(row=row, column=5, value=meta["label"])
            ws.cell(row=row, column=6, value=amt).number_format = "#,##0.00"
            ws.cell(row=row, column=7, value=cnt)
            row += 1
    auto_width(ws, [6, 18, 8, 10, 40, 12, 11])


def sheet_lifetime_vendors(wb, lifetime, naics_lookup):
    ws = wb.create_sheet("Lifetime_Vendors")
    title_row(ws, 1, 1, "Top 100 vendors across the window (in-scope, MIB stripped)")
    ws.cell(row=2, column=1, value="Sorted by lifetime $. Parent-UEI rolled. NAICS via SAM Entity Management API.").font = Font(italic=True, color="666666")

    uei_to_naics = {}
    for r in naics_lookup:
        uei_to_naics[r["uei"]] = r

    header_row(ws, 4, ["Rank", "Vendor", "Parent UEI", "Foreign?", "NAICS",
                        "Industry", "Sector", "Lifetime $M", "# records",
                        "# FYs", "# PIIDs", "Active FYs", "PIIDs"])
    row = 5
    for r in lifetime[:100]:
        n = uei_to_naics.get(r["uei"], {})
        ws.cell(row=row, column=1, value=int(r["rank"]))
        ws.cell(row=row, column=2, value=r["vendor"][:60])
        ws.cell(row=row, column=3, value=r["uei"]).font = Font(name="Menlo", size=10)
        ws.cell(row=row, column=4, value=r["foreign"])
        ws.cell(row=row, column=5, value=n.get("primary_naics", "")).font = Font(name="Menlo", size=10)
        ws.cell(row=row, column=6, value=n.get("naics_desc", "")[:50])
        ws.cell(row=row, column=7, value=n.get("naics_sector_label", ""))
        ws.cell(row=row, column=8, value=float(r["amount_M_lifetime"])).number_format = "#,##0.00"
        ws.cell(row=row, column=9, value=int(r["records"]))
        ws.cell(row=row, column=10, value=int(r["fy_count"]))
        ws.cell(row=row, column=11, value=int(r["piid_count"]))
        ws.cell(row=row, column=12, value=r["fys"])
        ws.cell(row=row, column=13, value=r["piids"])
        row += 1
    auto_width(ws, [6, 50, 16, 10, 10, 40, 26, 14, 10, 8, 9, 30, 30])


def sheet_naics_lookup(wb, naics_lookup):
    ws = wb.create_sheet("NAICS_Lookup")
    title_row(ws, 1, 1, "UEI → primary NAICS audit trail")
    ws.cell(row=2, column=1, value="Result from SAM Entity Management API for the top-N parent UEIs by lifetime $. Cached to sam_entity_lookups/.").font = Font(italic=True, color="666666")

    header_row(ws, 4, ["Rank", "Parent UEI", "Vendor", "Lifetime $M",
                        "Primary NAICS", "Industry", "NAICS 2-digit",
                        "NAICS 4-digit", "Sector", "CAGE", "Country", "Status"])
    row = 5
    for r in naics_lookup:
        ws.cell(row=row, column=1, value=int(r["rank"]) if r.get("rank") else "")
        ws.cell(row=row, column=2, value=r["uei"]).font = Font(name="Menlo", size=10)
        ws.cell(row=row, column=3, value=r["vendor"][:60])
        ws.cell(row=row, column=4, value=float(r["amount_M_lifetime"])).number_format = "#,##0.00"
        ws.cell(row=row, column=5, value=r["primary_naics"]).font = Font(name="Menlo", size=10)
        ws.cell(row=row, column=6, value=r["naics_desc"][:60])
        ws.cell(row=row, column=7, value=r["naics_2digit"])
        ws.cell(row=row, column=8, value=r["naics_4digit"])
        ws.cell(row=row, column=9, value=r["naics_sector_label"])
        ws.cell(row=row, column=10, value=r["cage"]).font = Font(name="Menlo", size=10)
        ws.cell(row=row, column=11, value=r["country"])
        ws.cell(row=row, column=12, value=r["lookup_status"])
        if r["lookup_status"] != "ok":
            for c in range(1, 13):
                ws.cell(row=row, column=c).fill = WARN_FILL
        row += 1
    auto_width(ws, [6, 16, 50, 12, 12, 50, 12, 12, 26, 12, 10, 12])


def sheet_trends(wb, trends):
    """FY-over-FY trend table."""
    ws = wb.create_sheet("Trends")
    title_row(ws, 1, 1, "Year-over-year trends — visible subaward $, vendor concentration, work-type mix")
    ws.cell(row=2, column=1,
            value="In-scope new-construction PIIDs, MIB pass-throughs already excluded. "
                  "HHI scale: 0-1500 competitive, 1500-2500 moderately concentrated, >2500 highly concentrated. "
                  "Sub-base measured as count of unique parent UEIs receiving any $ that FY."
            ).font = Font(italic=True, color="666666")

    header_row(ws, 4, [
        "FY", "Visible subs $M", "YoY %", "# records",
        "# unique parent UEIs", "Top-5 share", "Top-10 share",
        "HHI (0-10000)", "Concentration", "Top 3 NAICS-4 industries",
    ])
    row = 5
    for r in trends:
        fy = int(r["fy"])
        ws.cell(row=row, column=1, value=fy)
        ws.cell(row=row, column=2, value=float(r["total_$M"])).number_format = "#,##0.0"
        if r["yoy_pct"]:
            ws.cell(row=row, column=3, value=float(r["yoy_pct"])).number_format = "+0.0%;-0.0%;—"
        ws.cell(row=row, column=4, value=int(r["records"]))
        ws.cell(row=row, column=5, value=int(r["unique_parent_ueis"]))
        ws.cell(row=row, column=6, value=float(r["top5_share"])).number_format = "0.0%"
        ws.cell(row=row, column=7, value=float(r["top10_share"])).number_format = "0.0%"
        ws.cell(row=row, column=8, value=float(r["hhi_0_to_10000"])).number_format = "#,##0"
        ws.cell(row=row, column=9, value=r["concentration_label"])
        ws.cell(row=row, column=10, value=r["top3_naics4_$"])
        # Highlight FY25-26 (lag-depressed)
        if fy >= 2025:
            for c in range(1, 11):
                ws.cell(row=row, column=c).fill = WARN_FILL
        row += 1

    # Narrative annotation
    row += 2
    title_row(ws, row, 1, "Pattern in the data")
    row += 1
    notes = [
        "• Concentration: top-5 vendors held ~99% of $ in 2016-18 (a few primes dominated visible subs); fell to 31% by FY22 as the supplier base broadened.",
        "• Vendor base: from ~1 unique parent UEI in FY16 to 371 in FY23 — a ~370× expansion of the visible first-tier supplier network.",
        "• FY23 inflection: $2,048M visible subs (vs ~$534M FY22) is the MIB-era inflection, even after stripping BlueForge. This is the supplier-network-build investment showing up in FFATA.",
        "• FY24 step-down: $1,300M — partly real (BlueForge already excluded), partly FFATA lag working through.",
        "• FY25/26 figures are heavily lag-depressed; expect upward revisions over 12-18 months.",
        "• NAICS-3364 (Aircraft Parts / where Northrop and Leonardo classify) dominates — but that's a corporate-NAICS artifact, not a literal claim about submarine work.",
    ]
    for n in notes:
        cell = ws.cell(row=row, column=1, value=n)
        cell.alignment = Alignment(wrap_text=True, vertical="top")
        row += 1

    auto_width(ws, [6, 14, 9, 10, 14, 11, 11, 12, 20, 60])


def sheet_geographic(wb, geo_state, geo_country, foreign_share):
    """Where the dollars go geographically."""
    ws = wb.create_sheet("Geographic")
    title_row(ws, 1, 1, "Geographic distribution of new-construction submarine subs")
    ws.cell(row=2, column=1,
            value="Vendor's registered physical address from SAM. Reflects parent-UEI location, "
                  "not necessarily where the work is performed (a CA-headquartered prime may "
                  "subcontract work done elsewhere). National-security industrial-base footprint signal."
            ).font = Font(italic=True, color="666666")

    # ---- US state table ----
    title_row(ws, 4, 1, "By US state (descending by $M)")
    header_row(ws, 6, ["State", "State name", "$M lifetime",
                        "% of US total", "# records", "# unique vendors"])
    row = 7
    for r in geo_state:
        ws.cell(row=row, column=1, value=r["state"])
        ws.cell(row=row, column=2, value=r["state_name"])
        ws.cell(row=row, column=3, value=float(r["amount_M"])).number_format = "#,##0.0"
        ws.cell(row=row, column=4, value=float(r["pct_of_us_total"])).number_format = "0.0%"
        ws.cell(row=row, column=5, value=int(r["records"]))
        ws.cell(row=row, column=6, value=int(r["unique_vendors"]))
        row += 1

    # ---- Country table ----
    row += 2
    title_row(ws, row, 1, "By country (descending by $M)")
    row += 2
    header_row(ws, row, ["Country code", "Country", "$M lifetime",
                          "% of total", "# records", "# unique vendors"])
    row += 1
    for r in geo_country:
        ws.cell(row=row, column=1, value=r["country_code"]).font = Font(name="Menlo", size=11)
        ws.cell(row=row, column=2, value=r["country"])
        ws.cell(row=row, column=3, value=float(r["amount_M"])).number_format = "#,##0.0"
        ws.cell(row=row, column=4, value=float(r["pct_of_total"])).number_format = "0.00%"
        ws.cell(row=row, column=5, value=int(r["records"]))
        ws.cell(row=row, column=6, value=int(r["unique_vendors"]))
        if r["country_code"] != "US":
            for c in range(1, 7):
                ws.cell(row=row, column=c).fill = WARN_FILL
        row += 1

    # ---- Foreign share by FY ----
    row += 2
    title_row(ws, row, 1, "Foreign vendor share per FY")
    row += 2
    header_row(ws, row, ["FY", "Domestic $M", "Foreign $M", "Total $M", "Foreign share"])
    row += 1
    for r in foreign_share:
        fy = int(r["fy"])
        ws.cell(row=row, column=1, value=fy)
        ws.cell(row=row, column=2, value=float(r["domestic_$M"])).number_format = "#,##0.0"
        ws.cell(row=row, column=3, value=float(r["foreign_$M"])).number_format = "#,##0.0"
        ws.cell(row=row, column=4, value=float(r["total_$M"])).number_format = "#,##0.0"
        ws.cell(row=row, column=5, value=float(r["foreign_share"])).number_format = "0.0%"
        row += 1

    # Narrative
    row += 2
    title_row(ws, row, 1, "Reading the geography")
    row += 1
    notes = [
        "• 96.7% of visible subs $ flows to US-registered vendors. Foreign share is small (~3%) but concentrated in UK (BAE / Rosyth Royal Dockyard / Goodwin Steel Castings) and Switzerland (likely components).",
        "• California dominates (26% of US total) because Northrop Grumman's defense electronics segment is registered there. This overstates physical work done in CA — much of NG's submarine-related work is delivered from sites in other states.",
        "• Pennsylvania, Massachusetts, Wisconsin, New Jersey, New York are the next tier — strong manufacturing/fabrication ecosystem.",
        "• Connecticut shows only $196M (3% of US) — surprising given GDEB is HQ'd there. Reason: subs to GDEB are upstream of GDEB, not in CT. GDEB itself isn't a 'recipient' in this data — it's the prime.",
        "• Virginia at $185M (3%) is HII-NNS work flowing back as sub (under-counts the team-build share — see HII_NNS_Context tab).",
        "• Geographic concentration risk: top 5 states (CA, PA, MA, WI, NJ) = 60% of US-vendor $.",
    ]
    for n in notes:
        cell = ws.cell(row=row, column=1, value=n)
        cell.alignment = Alignment(wrap_text=True, vertical="top")
        row += 1

    auto_width(ws, [10, 22, 14, 14, 12, 14])


def sheet_hii_context(wb, hii_context):
    """HII Newport News segment revenue and the FFATA visibility gap."""
    ws = wb.create_sheet("HII_NNS_Context")
    title_row(ws, 1, 1, "HII Newport News segment context — the workshare federal data doesn't see")
    ws.cell(row=2, column=1,
            value="NNS segment revenue from HII 10-K filings (SEC EDGAR, structured). "
                  "Implied submarine portion uses analyst-consensus 25-35% share of NNS — NOT from HII filings. "
                  "Compared to HII-as-GDEB-sub flow we see in FFATA. The gap is the team-build workshare."
            ).font = Font(italic=True, color="666666")

    header_row(ws, 4, [
        "FY", "NNS seg rev $M", "NNS OpInc $M", "NNS op margin",
        "Implied sub $M (25%)", "Implied sub $M (30%)", "Implied sub $M (35%)",
        "FFATA visible HII-as-sub $M", "Gap ratio (30%/FFATA)", "10-K source",
    ])
    row = 5
    for r in hii_context:
        fy = int(r["fy"])
        ws.cell(row=row, column=1, value=fy)
        ws.cell(row=row, column=2, value=float(r["nns_segment_rev_$M"])).number_format = "#,##0.0"
        ws.cell(row=row, column=3, value=float(r["nns_seg_op_income_$M"])).number_format = "#,##0.0"
        ws.cell(row=row, column=4, value=float(r["nns_op_margin_pct"])).number_format = "0.0%"
        ws.cell(row=row, column=5, value=float(r["implied_sub_rev_low_25pct_$M"])).number_format = "#,##0.0"
        ws.cell(row=row, column=6, value=float(r["implied_sub_rev_mid_30pct_$M"])).number_format = "#,##0.0"
        ws.cell(row=row, column=7, value=float(r["implied_sub_rev_high_35pct_$M"])).number_format = "#,##0.0"
        ws.cell(row=row, column=8, value=float(r["ffata_visible_hii_as_sub_$M"])).number_format = "#,##0.000"
        if r["gap_ratio_30pct_vs_ffata"]:
            ws.cell(row=row, column=9, value=float(r["gap_ratio_30pct_vs_ffata"])).number_format = "#,##0×"
        else:
            ws.cell(row=row, column=9, value="∞ (no FFATA)")
        ws.cell(row=row, column=10, value=r["source_book"])
        row += 1

    # Narrative
    row += 2
    title_row(ws, row, 1, "Interpretation")
    row += 1
    notes = [
        "• NNS segment revenue is the structured top-line: ~$5.6-6.5B/yr through FY22-FY25. This is verified from HII 10-K filings (XBRL-structured, audited).",
        "• HII never discloses the submarine portion of NNS revenue. Defense analyst estimates put submarine work at 25-35% of NNS, with the balance being aircraft carrier construction (CVN 80/81/82), RCOH refueling (Stennis), and submarine fleet support services.",
        "• At a 30% midpoint: implied submarine revenue at HII-NNS is ~$1.5-2.0B/yr.",
        "• But FFATA-visible HII-as-GDEB-sub is essentially zero in our data — only $15.6M visible in FY21, and ~$0 in FY22-FY25. The gap ratio is 100-9,800×.",
        "• Why the gap? HII receives Virginia workshare under teaming-agreement terms where GDEB is vendor-of-record on FPDS, not as a discrete FFATA-reportable subaward. The workshare flows through accounting mechanisms that don't show up in FFATA.",
        "• Implication: federal first-tier subaward data captures ~5-10% of the true 'outsourced from GDEB' picture when HII is included. The FY_Headline ratio of ~15-30% is therefore a floor, not a ceiling.",
        "• What you'd add for a fuller picture: HII implied submarine revenue (1.5-2.0B/yr) PLUS the FY_Headline visible-subs total (~$1-2B/yr in FY23-FY25) → real annual 'outsourced from the Virginia prime team' figure is in the $2.5-4B range during FY23-FY25.",
    ]
    for n in notes:
        cell = ws.cell(row=row, column=1, value=n)
        cell.alignment = Alignment(wrap_text=True, vertical="top")
        row += 1

    auto_width(ws, [6, 14, 13, 11, 17, 17, 17, 18, 18, 14])


def sheet_ceo_commentary(wb, quotes):
    """Curated CEO/CFO quotes from earnings call transcripts."""
    ws = wb.create_sheet("CEO_Commentary")
    title_row(ws, 1, 1, "HII CEO/CFO commentary — curated quotes from earnings call transcripts")
    ws.cell(row=2, column=1,
            value="Transcripts scraped from Motley Fool, Insider Monkey, Yahoo Finance "
                  "(FY22-FY26 coverage; see hii_earnings_transcripts/README.md). "
                  "Quotes selected by topic-keyword + signal-strength scoring (presence of dollar "
                  "figures, percentages, specific numbers, forward-looking language)."
            ).font = Font(italic=True, color="666666")

    # Group by topic
    topics = {}
    for q in quotes:
        topics.setdefault(q["topic"], []).append(q)

    row = 4
    for topic, topic_quotes in topics.items():
        title_row(ws, row, 1, topic)
        row += 1
        header_row(ws, row, ["FY", "Q", "Speaker", "Quote"])
        row += 1
        for q in sorted(topic_quotes, key=lambda x: (int(x["fy"]), int(x["quarter"]))):
            ws.cell(row=row, column=1, value=int(q["fy"]))
            ws.cell(row=row, column=2, value=int(q["quarter"]))
            ws.cell(row=row, column=3, value=q["speaker"])
            cell = ws.cell(row=row, column=4, value=q["quote"])
            cell.alignment = Alignment(wrap_text=True, vertical="top")
            row += 1
        row += 1

    auto_width(ws, [6, 4, 22, 130])
    # Make rows taller for the quote cells
    for r_idx in range(4, row + 1):
        ws.row_dimensions[r_idx].height = None  # auto, but wrap_text will expand


def sheet_caveats(wb):
    ws = wb.create_sheet("Caveats")
    title_row(ws, 1, 1, "Caveats & methodology notes")

    sections = [
        ("Scope decisions", [
            "• Excluded N0002420C4312 (Hartford SSN-768 EOH) — this is a mid-life Engineered Overhaul of an existing in-service boat, not new construction.",
            "• Excluded N0002419C2125 (Va Tech Instructions / HPAD backfit) — upgrades/modifications to existing boats, not new build.",
            "• Excluded MIB / workforce pass-throughs by parent UEI: BlueForge Alliance (F8PEZKXES8B1), Training Modernization Group (QLJZVM6XKR71), Institute for Advanced Learning and Research (TCM3R4JPRKY4). These are workforce/supplier-infrastructure consortium flows, not construction subawards.",
            "• Included BPMI prime PIIDs (N0002419C2114, N0002419C2115, N0002424C2114) — naval reactor components delivered as GFE for new boats, even though BPMI is a separately-contracted prime, not a GDEB sub.",
            "• Included other GFE primes (LM combat systems, BAE forward subassembly, Rolls-Royce rotor) — these are non-GDEB primes delivering new-construction components.",
        ]),
        ("HII-NNS team-build visibility gap", [
            "• HII-NNS (Huntington Ingalls Newport News Shipbuilding) holds ~50% of Virginia + ~22% of Columbia construction work under teaming agreements where GDEB is vendor-of-record.",
            "• This work flows from Navy → GDEB → HII-NNS, but FFATA reports it as a GDEB subaward to HII only when HII actually accepts a sub PO from GDEB. Most of the workshare flows through other mechanisms that do not show up here.",
            "• Visible HII-NNS subs in our data total only ~$98M lifetime vs an estimated $4-6B/yr true team-build flow. This is the single biggest under-count in this workbook.",
            "• April 2025 NAVSEA announcement of a ~$1.3B FY24 Virginia mod naming HII-NNS directly does not appear here, because FPDS records GDEB as the vendor of record.",
        ]),
        ("FFATA reporting lag", [
            "• First-tier subaward filings lag prime contract activity by 6-18 months.",
            "• FY25 figures are heavily lag-depressed (~50-70% complete). Expect them to climb 2-4× over the following year.",
            "• FY26 figures are essentially empty (~0-5% complete). Do not draw any conclusions from FY26 cells.",
            "• FY20-FY21 may also be incomplete due to FFATA system maturation in that era.",
        ]),
        ("Basic Construction denominator", [
            "• Source: SCN P-5c cost-category line for 'Basic Construction/Conversion' on LI 1045 (Columbia) and LI 2013 (Virginia).",
            "• Reconciled across PB22-PB27 Navy budget books — for each FY, uses the most-revised actual from the most-recent book in which the FY appears as 'actual'.",
            "• Columbia is procured on a less-frequent cadence (FY21, FY24, FY26, FY27 in our window). 'No Col procurement' FYs are not missing data — they're intentional zeros.",
            "• Pre-FY22 Basic Construction is unavailable (would need PB20/PB21 books).",
            "• Basic Construction is per-ship-at-authorization, NOT per-FY cash flow. Per-FY ratios are lumpy; cumulative ratios are cleaner.",
        ]),
        ("NAICS work-type axis", [
            "• NAICS is a vendor's PRIMARY self-reported industry classifier from SAM.gov entity registration. It reflects WHAT THE FIRM IS (their main line of business), not necessarily what a specific sub-action was for.",
            "• Top-150 vendors by lifetime $ were enriched via SAM Entity Management API. Long-tail vendors not enriched are bucketed as 'UNENR'.",
            "• 4-digit NAICS (industry group) is used for FY × work-type aggregation — coarse enough to keep tables readable, fine enough to discriminate fabrication from electronics from engineering services.",
            "• Some firms list a primary NAICS that does not match their submarine work (e.g., a generic 'engineering services' code). NAICS quality is uneven and varies by firm.",
        ]),
        ("Dollar treatment", [
            "• subAwardAmount is summed by record. SAM data is verified dedup-clean (every subAwardReportId is unique).",
            "• FY attribution is by subAwardDate (action date), not submittedDate (filing date).",
            "• Values are nominal $ (not inflation-adjusted).",
            "• Some records have negative subAwardAmount — these are downward modifications and are kept (they net out correctly in sums).",
        ]),
        ("Limits of FFATA visibility", [
            "• FFATA only requires first-tier subaward reporting. Second-tier and below are invisible.",
            "• Some primes systematically under-report. GDEB does report; some other primes (notably HII-NNS via teaming) do not.",
            "• Federal naval shipyards (Norfolk, Portsmouth, Pearl Harbor, Puget Sound) execute with federal payroll, not subawards — not in scope, not in our data.",
        ]),
    ]
    row = 3
    for header, lines in sections:
        cell = ws.cell(row=row, column=1, value=header)
        cell.font = Font(bold=True, size=12, color="1F3864")
        row += 1
        for line in lines:
            c = ws.cell(row=row, column=1, value=line)
            c.alignment = Alignment(wrap_text=True, vertical="top")
            row += 1
        row += 1
    ws.column_dimensions["A"].width = 140


def main():
    scope = load_scope()
    by_piid = load_by_piid()
    by_vendor = load_by_vendor()
    lifetime = load_lifetime()
    naics_lookup = load_naics()
    basic_construction = load_basic_construction()
    trends = load_trends()
    geo_state = load_geo_state()
    geo_country = load_geo_country()
    foreign_share = load_foreign_share_by_fy()
    hii_context = load_hii_context()
    curated_quotes = load_curated_quotes()

    wb = Workbook()
    # Remove default sheet
    default = wb.active
    wb.remove(default)

    sheet_cover(wb, scope)
    sheet_scope_piids(wb, scope)
    sheet_fy_headline(wb, by_piid, basic_construction)
    sheet_trends(wb, trends)
    sheet_fy_by_naics(wb, by_vendor, naics_lookup)
    sheet_fy_by_vendor(wb, by_vendor, naics_lookup)
    sheet_fy_by_piid(wb, by_piid, scope)
    sheet_geographic(wb, geo_state, geo_country, foreign_share)
    sheet_lifetime_vendors(wb, lifetime, naics_lookup)
    sheet_hii_context(wb, hii_context)
    sheet_ceo_commentary(wb, curated_quotes)
    sheet_naics_lookup(wb, naics_lookup)
    sheet_caveats(wb)

    wb.save(OUT_XLSX)
    print(f"\n✓ Wrote {OUT_XLSX}")
    print(f"  Sheets: {wb.sheetnames}")


if __name__ == "__main__":
    main()
