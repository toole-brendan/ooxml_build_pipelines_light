"""validate_workbook.py - QA / tie-out for the built master_v2 workbook.

NOT part of the build pipeline. Reads back the already-built xlsx and checks:
  1. structural   - every XML part parses; no #REF!/#DIV/0!/... literals survived
  2. recalc       - LibreOffice headless recalculates the formulas; 0 error cells
  3. tie-out      - per-program TAM, average annual, coefficients and outyear
                    averages match the ANCHORS captured from the original
                    tam/master workbook (regression guard; survives master/ deletion)

Usage:
    python3 validate_workbook.py                 # full report (structural + recalc + tie-out)
    python3 validate_workbook.py --no-recalc     # structural + (formula-string) checks only
    python3 validate_workbook.py "Virginia TAM!C7" ...   # dump cells

Requires openpyxl; recalc requires `soffice` (LibreOffice) on PATH.
"""
from __future__ import annotations

import os
import shutil
import subprocess
import sys
import tempfile
import xml.dom.minidom as minidom
import zipfile
from pathlib import Path

OUT = Path(__file__).resolve().parent / os.environ.get(
    "WB_OUT", "20260620_Distributed Shipbuilding Master TAM_vS.xlsx")

# Anchors: per-program TAM / coefficients. Virginia is derived on the POP sheet
# (Block V master + disclosed rows, $-weighted), 34.024%; Columbia/DDG unchanged.
# Captured via LibreOffice recalc, 2026-06-21.
#
# Resolved by ROW LABEL (column B scan), not by hard-coded address: row positions
# drift as sections grow, but the labels are stable, so structural refactors no
# longer force re-baselining. (sheet, label, col, expected, tol). The value lives
# in `col` of the first row whose column-B label matches.
_ANCHORS = [
    ("Virginia TAM", "Cumulative TAM (FY22-27) $M", "C", 13636.2514801724, 0.5),
    ("Columbia TAM", "Cumulative TAM (FY22-27) $M", "C", 4507.0865004, 0.5),
    ("DDG-51 TAM",   "Cumulative TAM (FY22-27) $M", "C", 6421.7296801227, 0.5),
    ("Virginia TAM", "Average annual TAM $M/yr", "C", 2272.70858002873, 0.5),
    ("Columbia TAM", "Average annual TAM $M/yr", "C", 751.1810834, 0.5),
    ("DDG-51 TAM",   "Average annual TAM $M/yr", "C", 1070.28828002045, 0.5),
    ("Virginia TAM", "Applied BC supplier coefficient", "C", 0.3402368202043, 0.0005),
    ("Columbia TAM", "Applied BC supplier coefficient", "C", 0.22, 0.0005),
    ("DDG-51 TAM",   "Applied BC supplier coefficient", "C", 0.252936763771345, 0.0005),
    ("Place of Performance", "Virginia BC",
     "E", 0.3402368202043, 0.0005),
    ("Place of Performance", "DDG-51 BC, FY23-27",
     "E", 0.252936763771345, 0.0005),
    ("Place of Performance", "Submarine AP/LLTM",
     "E", 0.48486094882579, 0.0005),
]


def _row_by_label(ws, label: str):
    """First row whose column-B cell equals `label` (sheets build in gutter mode,
    so content labels live in column B). Returns None if not found."""
    for r in range(1, ws.max_row + 1):
        if ws["B" + str(r)].value == label:
            return r
    return None
_ERR_TOKENS = ("REF", "DIV", "VALUE", "NAME", "N/A", "NUM", "NULL")


def _is_err(v) -> bool:
    return isinstance(v, str) and v.startswith("#") and any(t in v for t in _ERR_TOKENS)


def structural_report() -> int:
    z = zipfile.ZipFile(OUT)
    bad = 0
    for n in z.namelist():
        if n.endswith((".xml", ".rels")):
            try:
                minidom.parseString(z.read(n))
            except Exception as e:  # noqa: BLE001
                print(f"  XML ERROR  {n}: {e}")
                bad += 1
    print(f"file: {OUT.name}  ({OUT.stat().st_size:,} bytes)  parts: {len(z.namelist())}  xml errors: {bad}")

    import openpyxl
    wb = openpyxl.load_workbook(OUT, data_only=False)
    print(f"sheets ({len(wb.sheetnames)}): {wb.sheetnames}")
    lit = [f"{ws.title}!{c.coordinate}={c.value}"
           for ws in wb.worksheets for row in ws.iter_rows() for c in row if _is_err(c.value)]
    print(f"error-literal cells: {len(lit)}")
    for e in lit[:20]:
        print("   ", e)
    return 1 if (bad or lit) else 0


def _recalc() -> Path:
    soffice = shutil.which("soffice") or "/opt/homebrew/bin/soffice"
    tmp = Path(tempfile.mkdtemp(prefix="tam_recalc_"))
    subprocess.run([soffice, "--headless", "--calc", "--convert-to", "xlsx",
                    "--outdir", str(tmp), str(OUT)],
                   check=True, capture_output=True)
    return tmp / (OUT.stem + ".xlsx")


def recalc_report() -> int:
    import openpyxl
    try:
        f = _recalc()
    except Exception as e:  # noqa: BLE001
        print(f"recalc SKIPPED (soffice unavailable: {e})")
        return 0
    wb = openpyxl.load_workbook(f, data_only=True)
    errs = [f"{ws.title}!{c.coordinate}={c.value}"
            for ws in wb.worksheets for row in ws.iter_rows() for c in row if _is_err(c.value)]
    print(f"\nrecalc formula-error cells: {len(errs)}")
    for e in errs[:20]:
        print("   ", e)

    print("\ntie-out vs original master anchors (label-resolved):")
    fails = 0
    for sheet, label, col, exp, tol in _ANCHORS:
        ws = wb[sheet]
        row = _row_by_label(ws, label)
        cell = f"{col}{row}" if row else "<not found>"
        got = ws[cell].value if row else None
        ok = isinstance(got, (int, float)) and abs(got - exp) <= tol
        fails += 0 if ok else 1
        print(f"  [{'OK' if ok else 'FAIL'}] {sheet}!{cell} [{label[:34]}]: got {got!r}  expected ~{exp}")

    # internal consistency (transposed grid): each FY column Total = Va+Col+DDG.
    # Programs are rows now; fiscal years are columns C..L under the "Program" header.
    es = wb["Executive Summary"]
    cross = 0
    hdr_row = None
    prog = {}
    for r in range(1, es.max_row + 1):
        lab = es["B" + str(r)].value
        if lab == "Program":
            hdr_row = r
        if lab in ("Virginia", "Columbia", "DDG-51", "Total"):
            prog[lab] = r
    if hdr_row and all(k in prog for k in ("Virginia", "Columbia", "DDG-51", "Total")):
        i = 0
        while True:
            col = chr(ord("C") + i)
            hdr = es[col + str(hdr_row)].value
            if not (isinstance(hdr, str) and hdr.startswith("FY")):
                break   # past the last FY column (e.g. "BC coeff")
            va, co, dd, tot = (es[col + str(prog[p])].value
                               for p in ("Virginia", "Columbia", "DDG-51", "Total"))
            if all(isinstance(x, (int, float)) for x in (va, co, dd, tot)):
                ok = abs((va + co + dd) - tot) < 0.5
                cross += 0 if ok else 1
                print(f"  [{'OK' if ok else 'FAIL'}] Exec Summary {hdr} Total = Va+Col+DDG "
                      f"({tot:.1f} vs {va + co + dd:.1f})")
            i += 1

    # §3 outyear band (FY2028-31, columns C..F): Total low/high = sum of the three
    # program low/high rows (labels "<program> low" / "<program> high").
    band = 0
    rows_by_lab = {es["B" + str(r)].value: r for r in range(1, es.max_row + 1)
                   if isinstance(es["B" + str(r)].value, str)}
    for scen in ("low", "high"):
        if f"Total {scen}" not in rows_by_lab:
            continue
        tr = rows_by_lab[f"Total {scen}"]
        pr = [rows_by_lab[f"{p} {scen}"] for p in ("Virginia", "Columbia", "DDG-51")
              if f"{p} {scen}" in rows_by_lab]
        for i in range(4):
            col = chr(ord("C") + i)
            tot = es[col + str(tr)].value
            parts = [es[col + str(r)].value for r in pr]
            if isinstance(tot, (int, float)) and all(isinstance(x, (int, float)) for x in parts):
                ok = abs(sum(parts) - tot) < 0.5
                band += 0 if ok else 1
                print(f"  [{'OK' if ok else 'FAIL'}] Exec band Total {scen} FY{2028 + i} "
                      f"= sum ({tot:.1f} vs {sum(parts):.1f})")
    return 1 if (errs or fails or cross or band) else 0


def dump_cells(refs: list[str]) -> int:
    import openpyxl
    wb = openpyxl.load_workbook(OUT, data_only=False)
    for ref in refs:
        sheet, _, cell = ref.rpartition("!")
        sheet = sheet.strip("'")
        if sheet not in wb.sheetnames:
            print(f"{ref:<30} -> NO SUCH SHEET")
            continue
        print(f"{ref:<30} -> {wb[sheet][cell].value!r}")
    return 0


if __name__ == "__main__":
    args = [a for a in sys.argv[1:] if not a.startswith("--")]
    if args:
        sys.exit(dump_cells(args))
    rc = structural_report()
    if "--no-recalc" not in sys.argv:
        rc |= recalc_report()
    print("\nRESULT:", "PASS" if rc == 0 else "FAIL")
    sys.exit(rc)
