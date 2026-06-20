"""QA / validation tool for the built workbook.

NOT part of the build pipeline. The pipeline (workbook_ddg/) is strictly
stdlib-only raw OOXML; this script uses openpyxl purely to read back the
*already-built* file and sanity-check it. Never import this from a sheet
module or from lib.py.

Usage:
    python validate_workbook.py                       # structural report
    python validate_workbook.py "Inputs!D6" "Funnel!C6" ...   # dump cells

openpyxl does not evaluate formulas (only Excel/LibreOffice do), so cell
dumps show the formula string for formula cells and the literal for
hardcoded cells - which is exactly what we want to audit the build.
"""
from __future__ import annotations

import sys
import xml.dom.minidom as minidom
import zipfile
from pathlib import Path

# Built workbook lands at the project root (projects/distributed_shipbuilding/ddg/), one level up from here.
OUT = Path(__file__).resolve().parent / "20260601_Distributed Shipbuilding DDG_vS.xlsx"


def structural_report() -> int:
    z = zipfile.ZipFile(OUT)
    bad = 0
    for n in z.namelist():
        if n.endswith((".xml", ".rels")):
            try:
                minidom.parseString(z.read(n))
            except Exception as e:  # noqa: BLE001
                print(f"  XML ERROR  {n}: {e}")
                bad += 1
    print(f"file: {OUT.name}  ({OUT.stat().st_size:,} bytes)")
    print(f"parts: {len(z.namelist())}  | xml errors: {bad}")

    import openpyxl

    wb = openpyxl.load_workbook(OUT, data_only=False)
    print(f"sheets ({len(wb.sheetnames)}): {wb.sheetnames}")
    # Scan every cell for error literals that survived the build.
    errors: list[str] = []
    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for c in row:
                if isinstance(c.value, str) and c.value.startswith("#") and c.value.endswith("!"):
                    errors.append(f"{ws.title}!{c.coordinate} = {c.value}")
    print(f"error-literal cells: {len(errors)}")
    for e in errors[:20]:
        print("   ", e)
    return 1 if (bad or errors) else 0


def dump_cells(refs: list[str]) -> int:
    import openpyxl

    wb = openpyxl.load_workbook(OUT, data_only=False)
    for ref in refs:
        sheet, _, cell = ref.rpartition("!")
        sheet = sheet.strip("'")
        if sheet not in wb.sheetnames:
            print(f"{ref:<28} -> NO SUCH SHEET (have {wb.sheetnames})")
            continue
        v = wb[sheet][cell].value
        print(f"{ref:<28} -> {v!r}")
    return 0


if __name__ == "__main__":
    if len(sys.argv) > 1:
        sys.exit(dump_cells(sys.argv[1:]))
    sys.exit(structural_report())
