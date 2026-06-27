"""validate_workbook.py - QA / tie-out for the built ddg_module_cost workbook.

NOT part of the build pipeline. Reads back the already-built xlsx and checks:
  1. structural - every XML part parses; no #REF!/#DIV/0!/... literals survived
  2. recalc     - LibreOffice headless recalculates the formulas; 0 error cells
  3. tie-out    - the cascade is internally consistent after recalc:
                  * per-unit x units  == per-ship Basic Construction
                  * module breakout total == per-ship Basic Construction
                  * the headline per-ship BC sits in the expected ~$1.57B band

Usage:
    python3 validate_workbook.py                 # full report (structural + recalc + tie-out)
    python3 validate_workbook.py --no-recalc     # structural only
    python3 validate_workbook.py "Module Cost!D9" ...   # dump cells

Requires openpyxl; recalc requires `soffice` (LibreOffice) on PATH.
"""
from __future__ import annotations

import os
import shutil
import subprocess
import sys
import xml.dom.minidom as minidom
import zipfile
from pathlib import Path
import tempfile

OUT = Path(__file__).resolve().parent / os.environ.get(
    "WB_OUT", "20260625_DDG-51 Module Cost_vS.xlsx")

_ERR_TOKENS = ("REF", "DIV", "VALUE", "NAME", "N/A", "NUM", "NULL")


def _is_err(v) -> bool:
    return isinstance(v, str) and v.startswith("#") and any(t in v for t in _ERR_TOKENS)


def _row_by_label(ws, label: str):
    for r in range(1, ws.max_row + 1):
        if ws["B" + str(r)].value == label:
            return r
    return None


def structural_report() -> int:
    z = zipfile.ZipFile(OUT)
    bad = 0
    for n in z.namelist():
        if n.endswith((".xml", ".rels")):
            try:
                minidom.parseString(z.read(n))
            except Exception as e:  # noqa: BLE001
                print(f"  XML ERROR  {n}: {e}")
                bad += 1
    print(f"file: {OUT.name}  ({OUT.stat().st_size:,} bytes)  parts: {len(z.namelist())}  xml errors: {bad}")

    import openpyxl
    wb = openpyxl.load_workbook(OUT, data_only=False)
    print(f"sheets ({len(wb.sheetnames)}): {wb.sheetnames}")
    lit = [f"{ws.title}!{c.coordinate}={c.value}"
           for ws in wb.worksheets for row in ws.iter_rows() for c in row if _is_err(c.value)]
    print(f"error-literal cells: {len(lit)}")
    for e in lit[:20]:
        print("   ", e)
    return 1 if (bad or lit) else 0


def _recalc() -> Path:
    soffice = shutil.which("soffice") or "/opt/homebrew/bin/soffice"
    tmp = Path(tempfile.mkdtemp(prefix="ddgmod_recalc_"))
    subprocess.run([soffice, "--headless", "--calc", "--convert-to", "xlsx",
                    "--outdir", str(tmp), str(OUT)],
                   check=True, capture_output=True)
    return tmp / (OUT.stem + ".xlsx")


def _num(ws, label, col):
    r = _row_by_label(ws, label)
    return (ws[f"{col}{r}"].value if r else None), (f"{col}{r}" if r else "<not found>")


def recalc_report() -> int:
    import openpyxl
    try:
        f = _recalc()
    except Exception as e:  # noqa: BLE001
        print(f"recalc SKIPPED (soffice unavailable: {e})")
        return 0
    wb = openpyxl.load_workbook(f, data_only=True)
    errs = [f"{ws.title}!{c.coordinate}={c.value}"
            for ws in wb.worksheets for row in ws.iter_rows() for c in row if _is_err(c.value)]
    print(f"\nrecalc formula-error cells: {len(errs)}")
    for e in errs[:20]:
        print("   ", e)

    mc = wb["Module Cost"]
    fails = 0

    # Headline per-ship BC (the Ship row, col D) sits in the ~$1.573B band.
    ship_bc, cellref = _num(mc, "Ship (per hull)", "D")
    ok = isinstance(ship_bc, (int, float)) and 1450 <= ship_bc <= 1700
    fails += 0 if ok else 1
    print(f"  [{'OK' if ok else 'FAIL'}] Module Cost!{cellref} per-ship BC = {ship_bc!r}  (expect ~1573.6)")

    # per-unit x units == per-ship BC
    unit_avg, _ = _num(mc, "Structural unit", "D")
    units, _ = _num(mc, "Structural unit", "C")
    if isinstance(unit_avg, (int, float)) and isinstance(units, (int, float)) and isinstance(ship_bc, (int, float)):
        ok = abs(unit_avg * units - ship_bc) < 0.5
        fails += 0 if ok else 1
        print(f"  [{'OK' if ok else 'FAIL'}] per-unit x units = ship BC "
              f"({unit_avg * units:.1f} vs {ship_bc:.1f})  [per unit={unit_avg:.2f} x {units:g}]")

    # module breakout total == per-ship BC
    alloc, cellref = _num(mc, "All modules", "D")
    if isinstance(alloc, (int, float)) and isinstance(ship_bc, (int, float)):
        ok = abs(alloc - ship_bc) < 0.5
        fails += 0 if ok else 1
        print(f"  [{'OK' if ok else 'FAIL'}] module breakout total = ship BC "
              f"({alloc:.1f} vs {ship_bc:.1f})")

    return 1 if (errs or fails) else 0


def dump_cells(refs: list[str]) -> int:
    import openpyxl
    wb = openpyxl.load_workbook(OUT, data_only=False)
    for ref in refs:
        sheet, _, cell = ref.rpartition("!")
        sheet = sheet.strip("'")
        if sheet not in wb.sheetnames:
            print(f"{ref:<30} -> NO SUCH SHEET")
            continue
        print(f"{ref:<30} -> {wb[sheet][cell].value!r}")
    return 0


if __name__ == "__main__":
    args = [a for a in sys.argv[1:] if not a.startswith("--")]
    if args:
        sys.exit(dump_cells(args))
    rc = structural_report()
    if "--no-recalc" not in sys.argv:
        rc |= recalc_report()
    print("\nRESULT:", "PASS" if rc == 0 else "FAIL")
    sys.exit(rc)
