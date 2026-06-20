"""extract_research_results - normalize the raw AI deep-research pulls into clean CSVs.

The raw pulls live (immutable) under research_pulls/<program>_subaward_research_completed.xlsx
- one workbook per program, each with an input worklist sheet plus a results sheet whose
name is inconsistent across files ("Research Results" / "Research Output"). This is the ONE
place that touches that raw format: it locates the results sheet by its header signature
(not its name), and writes a faithful, deterministic dump to

    extracted/<program>_subaward_research_results.csv
        Subawardee UEI | Role / Description | Source URLs | NAICS-6 code | NAICS-6 title

The dump is faithful on purpose - every value (including non-numeric NAICS codes like
"Not confirmed") is carried through so nothing from the pull is lost. All validation /
gap-fill precedence lives downstream in build_program_vendors (load_research_prose,
load_research_naics). Re-runnable; raw files are never modified.

Run: python3 scripts/extract_research_results.py
"""
from __future__ import annotations

import csv
from pathlib import Path

import openpyxl

REFACTOR = Path("/Users/brendantoole/projects3/ooxml_build_pipelines_light"
                "/projects/distributed_shipbuilding/research_shared/workbook_award_classification_refactor")
RAW_DIR = REFACTOR / "research_pulls"
EXTRACTED = REFACTOR / "extracted"

PROGRAMS = ["ddg", "virginia", "columbia"]
OUT_COLS = ["Subawardee UEI", "Role / Description", "Source URLs",
            "NAICS-6 code", "NAICS-6 title"]
# the results sheet is identified by these headers, wherever/however the sheet is named
SIG = {"Subawardee UEI", "Role / Description"}


def s(v) -> str:
    return "" if v is None else str(v).strip()


def find_results_sheet(wb) -> "openpyxl.worksheet.worksheet.Worksheet":
    for ws in wb.worksheets:
        headers = {s(c.value) for c in ws[1]}
        if SIG <= headers:
            return ws
    raise SystemExit(f"no results sheet (headers {SIG}) found in {wb}")


def extract(program: str) -> int:
    wb = openpyxl.load_workbook(RAW_DIR / f"{program}_subaward_research_completed.xlsx",
                               read_only=True, data_only=True)
    ws = find_results_sheet(wb)
    idx = {s(c.value): i for i, c in enumerate(ws[1])}
    src = ["Subawardee UEI", "Role / Description", "Source URLs",
           "NAICS-6 code", "NAICS-6 title"]
    out_rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        uei = s(row[idx["Subawardee UEI"]])
        if not uei:
            continue
        out_rows.append([s(row[idx[c]]) if c in idx else "" for c in src])
    out_path = EXTRACTED / f"{program}_subaward_research_results.csv"
    with out_path.open("w", encoding="utf-8", newline="") as fh:
        w = csv.writer(fh)
        w.writerow(OUT_COLS)
        w.writerows(out_rows)
    print(f"wrote {out_path.name}  ({len(out_rows)} rows)")
    return len(out_rows)


if __name__ == "__main__":
    for p in PROGRAMS:
        extract(p)
